"""
Outlook 草稿邮件管理器 - 增强版
支持从Excel读取数据，保存多个配置，预览并创建草稿邮件
新增功能：附件、BCC、批量创建、模板变量、配置导入导出等
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog
import json
import os
import win32com.client
import openpyxl
from datetime import datetime
import re
from pathlib import Path
import shutil
# Pillow for image rendering
try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
    from PIL import ImageGrab
except Exception:
    Image = None
    ImageDraw = None
    ImageFont = None
    ImageGrab = None
    ImageTk = None
import tempfile
import uuid
import threading
import time


class ScrollableFrame(ttk.Frame):
    """可滚动的Frame容器"""
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas_frame = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # 绑定Canvas大小变化，调整内部Frame宽度
        self.canvas.bind('<Configure>', self._on_canvas_configure)
        
        # 绑定鼠标滚轮 (仅当鼠标在区域内时)
        self.scrollable_frame.bind("<Enter>", self._bind_to_mousewheel)
        self.scrollable_frame.bind("<Leave>", self._unbind_from_mousewheel)

    def _on_canvas_configure(self, event):
        # 设置内部frame宽度等于canvas宽度
        self.canvas.itemconfig(self.canvas_frame, width=event.width)

    def _bind_to_mousewheel(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)

    def _unbind_from_mousewheel(self, event):
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        # 检查事件来源是否是内部可滚动控件（如 Text/ScrolledText）
        widget = event.widget
        while widget:
            if widget.__class__.__name__ in ('Text', 'ScrolledText'):
                return  # 让内部 Text 控件处理自己的滚动
            try:
                widget = widget.master
            except:
                break
        if event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")
        else:
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")


class ToolTip:
    """控件悬停提示工具类"""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.id = None
        self.x = self.y = 0
        
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)
        self.widget.bind("<ButtonPress>", self.leave)
        
    def enter(self, event=None):
        self.schedule()
        
    def leave(self, event=None):
        self.unschedule()
        self.hidetip()
        
    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(500, self.showtip)
        
    def unschedule(self):
        id = self.id
        self.id = None
        if id:
            self.widget.after_cancel(id)
            
    def showtip(self, event=None):
        try:
            x = y = 0
            x, y, cx, cy = self.widget.bbox("insert") # 对于Entry等控件获取光标位置
            x += self.widget.winfo_rootx() + 25
            y += self.widget.winfo_rooty() + 20
        except:
            # 对于Button等没有insert的控件
            x = self.widget.winfo_rootx() + 20
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        
        # 创建提示窗口
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry("+%d+%d" % (x, y))
        
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                       background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                       font=("Microsoft YaHei UI", "9", "normal"))
        label.pack(ipadx=2, ipady=1)
        
    def hidetip(self):
        tw = self.tip_window
        self.tip_window = None
        if tw:
            tw.destroy()


class OutlookDraftManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Outlook 草稿邮件管理器 - 增强版")
        self.root.geometry("1200x800")
        
        # 以脚本所在目录为基准路径，确保从任意目录启动都能找到配置文件
        self._base_dir = os.path.dirname(os.path.abspath(__file__))
        
        self.config_file = os.path.join(self._base_dir, "draft_configs.json")
        self.history_file = os.path.join(self._base_dir, "draft_history.json")
        self.configs = self.load_configs()
        self.current_excel_data = None
        self.excel_column_widths = []  # Excel 实际列宽（像素）
        self.attachments = []  # 附件列表
        self.inline_images = []  # 列表 of {'path':..., 'cid':...}
        self.excel_files = []  # 支持多个Excel文件
        self.selected_configs = []  # 批量处理选中的配置列表
        self.custom_placeholders = {}  # 自定义占位符字典 {'占位符名': '替换值'}
        self.email_signatures = self.load_signatures()  # 邮件签名字典
        self.content_templates = self.load_content_templates()  # 内容模板字典
        self.scheduled_sends = {}  # 定时发送字典 {草稿ID: 发送时间}
        self.current_language = 'zh'  # 当前语言 'zh' 或 'en'
        self.ui_scale = 1.0  # 界面缩放比例
        self.font_scale = 1.0  # 字体缩放比例
        
        # 新增：智能匹配相关变量
        self.smart_match_var = tk.BooleanVar(value=False)
        self.filename_keyword_var = tk.StringVar()

        # 自动保存定时器
        self.auto_save_timer = None
        
        # 拼写检查词典（常见错误）
        self.spell_check_dict = {
            # 英文常见错误
            'recieve': 'receive',
            'occured': 'occurred',
            'seperate': 'separate',
            'definately': 'definitely',
            'untill': 'until',
            'tommorrow': 'tomorrow',
            'sucessful': 'successful',
            'acheive': 'achieve',
            # 中文常见错误
            '以至': '以致',  # 语境相关
            '做': '作',  # 如"作为"
        }
        
        # 多语言文本
        self.translations = {
            'zh': {
                'title': 'Outlook 草稿邮件管理器 - 增强版',
                'config_name': '配置名称',
                'save_config': '保存配置',
                'load_config': '加载配置',
                'delete_config': '删除配置',
                'export_config': '导出配置',
                'import_config': '导入配置',
                'excel_source': 'Excel 数据源',
                'excel_file': 'Excel 文件',
                'select_file': '选择文件',
                'worksheet': '工作表',
                'data_range': '数据范围',
                'auto_detect': '自动检测',
                'read_data': '读取数据',
                'preview_data': '预览数据',
                'paste_formatting': '保留格式粘贴',
                'paste_picture': '粘贴为图片',
                'email_content': '邮件内容',
                'recipients': '收件人',
                'cc': '抄送',
                'bcc': '密送',
                'subject': '主题',
                'attachments': '附件',
                'body': '正文',
                'add': '添加',
                'delete': '删除',
                'clear': '清空',
                'create_draft': '创建草稿',
                'batch_create': '批量创建',
                'preview': '预览',
                'status': '就绪',
                'spell_check': '拼写检查',
                'schedule_send': '定时发送',
                'language': '语言',
            },
            'en': {
                'title': 'Outlook Draft Manager - Enhanced',
                'config_name': 'Config Name',
                'save_config': 'Save Config',
                'load_config': 'Load Config',
                'delete_config': 'Delete Config',
                'export_config': 'Export Config',
                'import_config': 'Import Config',
                'excel_source': 'Excel Data Source',
                'excel_file': 'Excel File',
                'select_file': 'Select File',
                'worksheet': 'Worksheet',
                'data_range': 'Data Range',
                'auto_detect': 'Auto Detect',
                'read_data': 'Read Data',
                'preview_data': 'Preview Data',
                'paste_formatting': 'Keep Formatting',
                'paste_picture': 'Paste as Picture',
                'email_content': 'Email Content',
                'recipients': 'To',
                'cc': 'CC',
                'bcc': 'BCC',
                'subject': 'Subject',
                'attachments': 'Attachments',
                'body': 'Body',
                'add': 'Add',
                'delete': 'Delete',
                'clear': 'Clear',
                'create_draft': 'Create Draft',
                'batch_create': 'Batch Create',
                'preview': 'Preview',
                'status': 'Ready',
                'spell_check': 'Spell Check',
                'schedule_send': 'Schedule Send',
                'language': 'Language',
            }
        }
        
        self.excel_history = []  # Excel 文件历史记录
        self.default_excel_folder = ""  # 默认 Excel 文件夹
        
        self.setup_ui()
        self.setup_keyboard_shortcuts()
        self.load_ui_preferences()  # 加载界面偏好设置
        self.start_auto_save()
        
    def setup_ui(self):
        """设置用户界面 - 优化版"""
        # 使用 Notebook 选项卡布局
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 创建可滚动的选项卡容器
        self.tab_config_wrapper = ScrollableFrame(self.notebook)
        self.tab_compose_wrapper = ScrollableFrame(self.notebook)
        
        self.notebook.add(self.tab_config_wrapper, text="⚙️ 设置与数据")
        self.notebook.add(self.tab_compose_wrapper, text="✉️ 邮件编辑")
        
        # 将实际内容放入可滚动区域
        self.tab_config = self.tab_config_wrapper.scrollable_frame
        self.tab_compose = self.tab_compose_wrapper.scrollable_frame
        
        # --- Tab 1: 设置与数据 ---
        self.setup_config_tab()
        
        # --- Tab 2: 邮件编辑 ---
        self.setup_compose_tab()
        
        # --- 底部操作栏 ---
        self.setup_bottom_bar()
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪 | 快捷键: Ctrl+S=保存配置 | Ctrl+D=创建草稿 | Ctrl+P=预览 | F1=帮助")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def setup_config_tab(self):
        """设置配置和数据选项卡"""
        # 使用 PanedWindow 分割左右 (左: 配置/Excel, 右: 配置浏览器)
        paned = ttk.PanedWindow(self.tab_config, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=3)
        
        # 右侧配置浏览器
        self.setup_config_browser(paned) 
        
        # --- 配置管理区域 ---
        config_frame = ttk.LabelFrame(left_frame, text="配置管理", padding="10")
        config_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 第一行：选择配置
        row1 = ttk.Frame(config_frame)
        row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="选择配置:").pack(side=tk.LEFT)
        self.config_combo = ttk.Combobox(row1, width=30, state="readonly")
        self.config_combo.pack(side=tk.LEFT, padx=5)
        self.config_combo.bind("<<ComboboxSelected>>", self.load_selected_config)
        self.update_config_list()
        
        ttk.Button(row1, text="新建", command=self.new_config).pack(side=tk.LEFT, padx=2)
        ttk.Button(row1, text="保存", command=self.save_current_config).pack(side=tk.LEFT, padx=2)
        ttk.Button(row1, text="删除", command=self.delete_config).pack(side=tk.LEFT, padx=2)
        
        # 第二行：配置名称与导入导出
        row2 = ttk.Frame(config_frame)
        row2.pack(fill=tk.X, pady=2)
        ttk.Label(row2, text="配置名称:").pack(side=tk.LEFT)
        self.config_name_var = tk.StringVar()
        ttk.Entry(row2, textvariable=self.config_name_var, width=25).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(row2, text="导出", command=self.export_config).pack(side=tk.LEFT, padx=2)
        ttk.Button(row2, text="导入", command=self.import_config).pack(side=tk.LEFT, padx=2)
        
        self.load_excel_path_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(row2, text="加载Excel路径", variable=self.load_excel_path_var).pack(side=tk.LEFT, padx=10)

        # 批量操作
        batch_frame = ttk.LabelFrame(left_frame, text="批量操作", padding="5")
        batch_frame.pack(fill=tk.X, padx=5, pady=5)
        
        btn_sel_multi = ttk.Button(batch_frame, text="📋 选择多项配置", command=self.select_multiple_configs)
        btn_sel_multi.pack(side=tk.LEFT, padx=5)
        ToolTip(btn_sel_multi, "勾选多个配置方案，一次性批量执行")
        
        btn_batch_run = ttk.Button(batch_frame, text="🚀 批量生成草稿", command=self.batch_create_all_drafts)
        btn_batch_run.pack(side=tk.LEFT, padx=5)
        ToolTip(btn_batch_run, "开始批量生成任务 (基于勾选的配置)")
        
        btn_batch_prev = ttk.Button(batch_frame, text="👁️ 批量预览", command=self.preview_all_configs)
        btn_batch_prev.pack(side=tk.LEFT, padx=5)
        ToolTip(btn_batch_prev, "检查所有选定配置的预览效果")
        
        # --- Excel 数据源区域 ---
        excel_frame = ttk.LabelFrame(left_frame, text="Excel 数据源", padding="10")
        excel_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 文件选择
        f_row = ttk.Frame(excel_frame)
        f_row.pack(fill=tk.X, pady=2)
        ttk.Label(f_row, text="Excel文件:").pack(side=tk.LEFT)
        self.excel_path_var = tk.StringVar()
        self.excel_path_combo = ttk.Combobox(f_row, textvariable=self.excel_path_var, state="normal", width=40)
        self.excel_path_combo.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.excel_path_combo.bind("<<ComboboxSelected>>", lambda e: self.load_sheets())
        ToolTip(self.excel_path_combo, "输入或选择Excel文件路径")
        
        btn_browse = ttk.Button(f_row, text="浏览文件...", command=self.select_excel_file)
        btn_browse.pack(side=tk.LEFT, padx=2)
        ToolTip(btn_browse, "打开文件选择对话框")
        
        # 移除旧的文件夹按钮，合并到"定位文件夹"
        # btn_folder = ttk.Button(f_row, text="浏览文件夹...", command=self.select_excel_folder)
        # btn_folder.pack(side=tk.LEFT, padx=2)
        
        btn_latest = ttk.Button(f_row, text="加载最新", command=self.load_latest_excel_file)
        btn_latest.pack(side=tk.LEFT, padx=2)
        ToolTip(btn_latest, "自动加载包含'出货计划'或其他关键词的最新Excel文件")
        
        btn_open_dir = ttk.Button(f_row, text="📂 定位文件夹", command=self.select_excel_folder)
        btn_open_dir.pack(side=tk.LEFT, padx=2)
        ToolTip(btn_open_dir, "指定数据源文件夹，并自动搜索其中最新的Excel文件")
        
        # 动态Excel选项
        dyn_row = ttk.Frame(excel_frame)
        dyn_row.pack(fill=tk.X, pady=2)
        self.prompt_excel_var = tk.BooleanVar(value=False)
        chk_prompt = ttk.Checkbutton(dyn_row, text="运行时询问", variable=self.prompt_excel_var)
        chk_prompt.pack(side=tk.LEFT, padx=5)
        ToolTip(chk_prompt, "勾选后，每次生成草稿时都会弹窗询问使用哪个Excel文件")
        
        chk_smart = ttk.Checkbutton(dyn_row, text="智能文件匹配", variable=self.smart_match_var)
        chk_smart.pack(side=tk.LEFT, padx=10)
        ToolTip(chk_smart, "勾选后，将自动查找最新版本的Excel文件 (支持跨文件夹)")

        self.attach_excel_var = tk.BooleanVar(value=False)
        chk_attach = ttk.Checkbutton(dyn_row, text="附带源文件", variable=self.attach_excel_var)
        chk_attach.pack(side=tk.LEFT, padx=10)
        ToolTip(chk_attach, "勾选后，生成草稿时会自动将当前使用的Excel文件作为附件")
        
        btn_smart = ttk.Button(dyn_row, text="🔍 智能查找工具", command=self.smart_search_excel)
        btn_smart.pack(side=tk.LEFT, padx=10)
        ToolTip(btn_smart, "手动配置智能查找规则")

        # 关键词行
        kw_row = ttk.Frame(excel_frame)
        kw_row.pack(fill=tk.X, pady=2)
        ttk.Label(kw_row, text="文件名关键词:").pack(side=tk.LEFT)
        ttk.Entry(kw_row, textvariable=self.filename_keyword_var, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Label(kw_row, text="(选填, 用于智能匹配时的筛选)", foreground="gray").pack(side=tk.LEFT)
        
        # 工作表与范围
        s_row = ttk.Frame(excel_frame)
        s_row.pack(fill=tk.X, pady=2)
        ttk.Label(s_row, text="工作表:").pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(s_row, width=15)
        self.sheet_combo.pack(side=tk.LEFT, padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)
        
        ttk.Label(s_row, text="范围:").pack(side=tk.LEFT, padx=(10,0))
        self.range_var = tk.StringVar(value="A1:C10")
        ttk.Entry(s_row, textvariable=self.range_var, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(s_row, text="自动检测", command=self.auto_detect_range).pack(side=tk.LEFT, padx=2)
        
        # 数据操作
        d_row = ttk.Frame(excel_frame)
        d_row.pack(fill=tk.X, pady=5)
        ttk.Button(d_row, text="读取数据", command=self.read_excel_data).pack(side=tk.LEFT, padx=2)
        ttk.Button(d_row, text="预览数据", command=self.preview_excel_data).pack(side=tk.LEFT, padx=2)
        ttk.Button(d_row, text="保留格式粘贴", command=self.paste_with_formatting).pack(side=tk.LEFT, padx=2)
        ttk.Button(d_row, text="粘贴为图片", command=self.paste_as_picture).pack(side=tk.LEFT, padx=2)
        
        # 列限制
        l_row = ttk.Frame(excel_frame)
        l_row.pack(fill=tk.X, pady=2)
        ttk.Label(l_row, text="检测列限制:").pack(side=tk.LEFT)
        self.col_limit_var = tk.StringVar()
        ttk.Entry(l_row, textvariable=self.col_limit_var, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Label(l_row, text="(可选, 例: A:C)", foreground="gray").pack(side=tk.LEFT)

        # 占位符应用（应用户要求加入第一页底部）
        ph_frame = ttk.LabelFrame(left_frame, text="占位符应用", padding="5")
        ph_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(ph_frame, text="📌 管理占位符", command=self.manage_placeholders).pack(side=tk.LEFT, padx=5)
        ttk.Label(ph_frame, text="可在所有场景（主题、正文、收件人）使用", foreground="gray", font=('TkDefaultFont', 8)).pack(side=tk.LEFT, padx=5)

    def setup_compose_tab(self):
        """设置邮件编辑选项卡"""
        frame = self.tab_compose
        
        # 收件人区域
        recipients_frame = ttk.LabelFrame(frame, text="收件人信息", padding="10")
        recipients_frame.pack(fill=tk.X, padx=5, pady=5)
        
        def create_recipient_row(parent, label_text, var):
            row = ttk.Frame(parent)
            row.pack(fill=tk.X, pady=2)
            ttk.Label(row, text=label_text, width=10).pack(side=tk.LEFT)
            entry = ttk.Entry(row, textvariable=var)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            # 添加 "+" 按钮用于插入占位符
            ttk.Button(row, text="+", width=3,
                     command=lambda: self.show_placeholder_menu(entry, is_recipient=True)).pack(side=tk.LEFT)
            return entry

        self.to_entry_var = tk.StringVar()
        create_recipient_row(recipients_frame, "收件人:", self.to_entry_var)
        
        self.cc_entry_var = tk.StringVar()
        create_recipient_row(recipients_frame, "抄送:", self.cc_entry_var)
        
        self.bcc_entry_var = tk.StringVar()
        create_recipient_row(recipients_frame, "密送:", self.bcc_entry_var)
        
        ttk.Label(recipients_frame, text="提示: 多个邮箱请用分号(;)分隔", foreground="gray", font=("", 8)).pack(anchor=tk.E)

        # 主题
        subject_frame = ttk.Frame(frame)
        subject_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(subject_frame, text="主题:", width=10).pack(side=tk.LEFT)
        self.subject_var = tk.StringVar()
        subject_entry = ttk.Entry(subject_frame, textvariable=self.subject_var)
        subject_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(subject_frame, text="插入占位符", command=lambda: self.show_placeholder_menu(subject_entry)).pack(side=tk.LEFT)
        
        # 正文
        body_frame = ttk.LabelFrame(frame, text="邮件正文 (支持HTML)", padding="5")
        body_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        toolbar = ttk.Frame(body_frame)
        toolbar.pack(fill=tk.X, pady=2)
        ttk.Button(toolbar, text="插入占位符", command=lambda: self.show_placeholder_menu(self.body_editor)).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="占位符管理", command=self.manage_placeholders).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="模板管理", command=self.manage_content_templates).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="签名管理", command=self.manage_signatures).pack(side=tk.LEFT, padx=2)
        
        self.body_editor = scrolledtext.ScrolledText(body_frame, height=15)
        self.body_editor.pack(fill=tk.BOTH, expand=True)
        # 添加格式化工具栏
        self.add_format_toolbar(body_frame, self.body_editor)
        # 兼容旧代码
        self.body_text = self.body_editor
        
        # 附件
        att_frame = ttk.LabelFrame(frame, text="附件 & 发送选项", padding="5")
        att_frame.pack(fill=tk.X, padx=5, pady=5)
        
        att_btn_row = ttk.Frame(att_frame)
        att_btn_row.pack(fill=tk.X, pady=2)
        
        # 附件按钮（增加Tooltip）
        btn_add = ttk.Button(att_btn_row, text="添加文件", command=self.add_attachment)
        btn_add.pack(side=tk.LEFT, padx=2)
        ToolTip(btn_add, "从电脑中选择文件作为附件")
        
        btn_excel = ttk.Button(att_btn_row, text="添加当前Excel", command=self.attach_current_excel)
        btn_excel.pack(side=tk.LEFT, padx=2)
        ToolTip(btn_excel, "将当前加载的Excel源文件作为附件")
        
        btn_del = ttk.Button(att_btn_row, text="删除选中", command=self.remove_selected_attachment)
        btn_del.pack(side=tk.LEFT, padx=2)
        
        btn_clear = ttk.Button(att_btn_row, text="清空", command=self.clear_attachments)
        btn_clear.pack(side=tk.LEFT, padx=2)
        
        # 发送选项
        opt_frame = ttk.Frame(att_frame)
        opt_frame.pack(fill=tk.X, pady=5)
        
        self.priority_var = tk.BooleanVar(value=False)
        chk_prio = ttk.Checkbutton(opt_frame, text="🔥 高优先级", variable=self.priority_var)
        chk_prio.pack(side=tk.LEFT, padx=10)
        ToolTip(chk_prio, "标记邮件为高重要性 (High Importance)")
        
        self.receipt_var = tk.BooleanVar(value=False)
        chk_receipt = ttk.Checkbutton(opt_frame, text="📫 已读回执", variable=self.receipt_var)
        chk_receipt.pack(side=tk.LEFT, padx=10)
        ToolTip(chk_receipt, "请求对方阅读后的回执 (Read Receipt)")
        
        self.attachment_listbox = tk.Listbox(att_frame, height=3)
        self.attachment_listbox.pack(fill=tk.X, pady=2)
        self.attachment_listbox.bind('<Double-Button-1>', self.open_selected_attachment)

    def setup_bottom_bar(self):
        """底部操作栏"""
        bar = ttk.Frame(self.root, padding="5")
        bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        left = ttk.Frame(bar)
        left.pack(side=tk.LEFT)
        ttk.Button(left, text="生成预览", command=self.generate_preview).pack(side=tk.LEFT, padx=5)
        ttk.Button(left, text="增强预览", command=self.show_enhanced_preview).pack(side=tk.LEFT, padx=5)
        
        # 中间缩放控件
        center = ttk.Frame(bar)
        center.pack(side=tk.LEFT, expand=True)
        ttk.Button(center, text="🔍−", width=3, command=self.zoom_out).pack(side=tk.LEFT, padx=1)
        self.zoom_label = ttk.Label(center, text="100%", width=5, anchor=tk.CENTER)
        self.zoom_label.pack(side=tk.LEFT, padx=2)
        ttk.Button(center, text="🔍+", width=3, command=self.zoom_in).pack(side=tk.LEFT, padx=1)
        ttk.Button(center, text="🌐", width=3, command=self.toggle_language).pack(side=tk.LEFT, padx=10)
        
        right = ttk.Frame(bar)
        right.pack(side=tk.RIGHT)
        ttk.Button(right, text="创建草稿 (Ctrl+D)", command=self.create_draft).pack(side=tk.LEFT, padx=5)
        ttk.Button(right, text="定时发送", command=self.schedule_send_dialog).pack(side=tk.LEFT, padx=5)
    
    def setup_keyboard_shortcuts(self):
        """设置键盘快捷键"""
        # Ctrl+S - 保存配置
        self.root.bind('<Control-s>', lambda e: self.save_current_config())
        self.root.bind('<Control-S>', lambda e: self.save_current_config())
        
        # Ctrl+D - 创建草稿
        self.root.bind('<Control-d>', lambda e: self.create_draft())
        self.root.bind('<Control-D>', lambda e: self.create_draft())
        
        # Ctrl+P - 生成预览
        self.root.bind('<Control-p>', lambda e: self.generate_preview())
        self.root.bind('<Control-P>', lambda e: self.generate_preview())
        
        # Ctrl+R - 读取Excel数据
        self.root.bind('<Control-r>', lambda e: self.read_excel_data())
        self.root.bind('<Control-R>', lambda e: self.read_excel_data())
        
        # Ctrl+B - 批量创建
        self.root.bind('<Control-b>', lambda e: self.batch_create_drafts())
        self.root.bind('<Control-B>', lambda e: self.batch_create_drafts())
        
        # F1 - 显示帮助
        self.root.bind('<F1>', lambda e: self.show_help())
        
        # Ctrl+Q - 退出程序
        self.root.bind('<Control-q>', lambda e: self.root.quit())
        self.root.bind('<Control-Q>', lambda e: self.root.quit())
        
        # Ctrl+Plus / Ctrl+Minus - 缩放
        self.root.bind('<Control-plus>', lambda e: self.zoom_in())
        self.root.bind('<Control-equal>', lambda e: self.zoom_in())  # 不按Shift的+键
        self.root.bind('<Control-minus>', lambda e: self.zoom_out())
        self.root.bind('<Control-0>', lambda e: self.reset_zoom())
        
        # 鼠标滚轮+Ctrl - 缩放
        self.root.bind('<Control-MouseWheel>', self.mouse_zoom)
        
    def show_help(self):
        """显示快捷键帮助"""
        help_text = """
═══════════════════════════════
    快捷键列表
═══════════════════════════════

📋 配置管理:
   Ctrl+S    保存当前配置
   
✉️ 邮件操作:
   Ctrl+D    创建草稿邮件
   Ctrl+B    批量创建草稿
   Ctrl+P    生成邮件预览
   
📊 Excel 操作:
   Ctrl+R    读取 Excel 数据

🔍 界面缩放:
   Ctrl++    放大界面
   Ctrl+-    缩小界面
   Ctrl+0    重置缩放（100%）
   Ctrl+滚轮  鼠标滚轮缩放
   
❓ 其他:
   F1        显示此帮助
   Ctrl+Q    退出程序
   
═══════════════════════════════
💡 提示: 所有快捷键在程序任何位置都可使用
═══════════════════════════════
        """
        messagebox.showinfo("快捷键帮助", help_text)
        
    def load_configs(self):
        """加载保存的配置"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                messagebox.showerror("错误", f"加载配置失败: {str(e)}")
                return {}
        return {}
    
    def save_configs(self):
        """保存所有配置到文件"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.configs, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror("错误", f"保存配置失败: {str(e)}")
            return False
    
    def load_ui_preferences(self):
        """加载界面偏好设置"""
        pref_file = os.path.join(self._base_dir, "ui_preferences.json")
        if os.path.exists(pref_file):
            try:
                with open(pref_file, 'r', encoding='utf-8') as f:
                    prefs = json.load(f)
                    self.ui_scale = prefs.get('ui_scale', 1.0)
                    self.current_language = prefs.get('language', 'zh')
                    self.excel_history = prefs.get('excel_history', [])
                    self.default_excel_folder = prefs.get('default_excel_folder', "")
                    
                    # 更新历史记录下拉框
                    if hasattr(self, 'excel_path_combo'):
                        self.excel_path_combo['values'] = self.excel_history
                    
                    # 应用缩放
                    if self.ui_scale != 1.0:
                        self.root.after(100, self.apply_zoom)  # 延迟应用，确保界面已加载
            except Exception as e:
                print(f"加载UI偏好设置失败: {e}")
    
    def save_ui_preferences(self):
        """保存界面偏好设置"""
        pref_file = os.path.join(self._base_dir, "ui_preferences.json")
        try:
            prefs = {
                'ui_scale': self.ui_scale,
                'language': self.current_language,
                'excel_history': self.excel_history,
                'default_excel_folder': self.default_excel_folder
            }
            with open(pref_file, 'w', encoding='utf-8') as f:
                json.dump(prefs, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存UI偏好设置失败: {e}")
    
    def update_config_list(self):
        """更新配置下拉列表"""
        config_names = list(self.configs.keys())
        self.config_combo['values'] = config_names
        if config_names:
            self.config_combo.current(0)
        
        # 同时更新配置浏览器
        if hasattr(self, 'config_listbox'):
            self.update_config_browser()
    
    def new_config(self):
        """创建新配置"""
        self.config_name_var.set("")
        self.excel_path_var.set("")
        self.range_var.set("A1:C10")
        
        # 清空收件人列表
        self.to_entry_var.set("")
        self.cc_entry_var.set("")
        self.bcc_entry_var.set("")
        
        self.subject_var.set("")
        self.body_text.delete(1.0, tk.END)
        self.sheet_combo.set("")
        self.current_excel_data = None
        self.attachments = []
        self.update_attachment_list()
        self.status_var.set("新建配置")
    
    def save_current_config(self):
        """保存当前配置"""
        config_name = self.config_name_var.get().strip()
        if not config_name:
            messagebox.showwarning("警告", "请输入配置名称")
            return
        
        # 获取所有邮箱
        to_list = [x.strip() for x in self.to_entry_var.get().split(';') if x.strip()]
        cc_list = [x.strip() for x in self.cc_entry_var.get().split(';') if x.strip()]
        bcc_list = [x.strip() for x in self.bcc_entry_var.get().split(';') if x.strip()]
        
        config_data = {
            "excel_path": self.excel_path_var.get(),
            "sheet_name": self.sheet_combo.get(),
            "data_range": self.range_var.get(),
            "to": to_list,  # 保存为列表
            "cc": cc_list,
            "bcc": bcc_list,
            "subject": self.subject_var.get(),
            "body": self.body_text.get(1.0, tk.END).strip(),
            "attachments": self.attachments.copy(),
            "inline_images": self.inline_images.copy() if hasattr(self, 'inline_images') else [], # 保存内嵌图片信息
            "custom_placeholders": self.custom_placeholders.copy(),  # 保存自定义占位符
            "high_priority": self.priority_var.get(), # 保存优先级
            "read_receipt": self.receipt_var.get(),    # 保存已读回执
            "smart_match": self.smart_match_var.get(), # 保存智能匹配开关
            "filename_keyword": self.filename_keyword_var.get(), # 保存关键词
            "attach_excel": self.attach_excel_var.get() # 保存是否附带源文件
        }
        
        self.configs[config_name] = config_data
        if self.save_configs():
            self.update_config_list()
            # 选中刚保存的配置
            index = list(self.configs.keys()).index(config_name)
            self.config_combo.current(index)
            messagebox.showinfo("成功", f"配置 '{config_name}' 已保存")
            self.status_var.set(f"已保存配置: {config_name}")
    
    def load_selected_config(self, event=None):
        """加载选中的配置"""
        config_name = self.config_combo.get()
        if not config_name or config_name not in self.configs:
            return
        
        config = self.configs[config_name]
        self.config_name_var.set(config_name)
        
        # 只有在勾选了"加载Excel路径"时才覆盖当前路径
        if self.load_excel_path_var.get():
            self.excel_path_var.set(config.get("excel_path", ""))
            
        self.range_var.set(config.get("data_range", "A1:C10"))
        
        # 加载收件人列表
        to_data = config.get("to", [])
        if isinstance(to_data, list):
            self.to_entry_var.set("; ".join(to_data))
        else:
            self.to_entry_var.set(str(to_data))
            
        cc_data = config.get("cc", [])
        if isinstance(cc_data, list):
            self.cc_entry_var.set("; ".join(cc_data))
        else:
            self.cc_entry_var.set(str(cc_data))
            
        bcc_data = config.get("bcc", [])
        if isinstance(bcc_data, list):
            self.bcc_entry_var.set("; ".join(bcc_data))
        else:
            self.bcc_entry_var.set(str(bcc_data))
        
        self.subject_var.set(config.get("subject", ""))
        self.body_text.delete(1.0, tk.END)
        self.body_text.insert(1.0, config.get("body", ""))
        self.attachments = config.get("attachments", [])
        self.inline_images = config.get("inline_images", []) # 加载内嵌图片信息
        self.update_attachment_list()
        
        # 加载发送选项
        self.priority_var.set(config.get("high_priority", False))
        self.receipt_var.set(config.get("read_receipt", False))
        self.smart_match_var.set(config.get("smart_match", False))
        self.filename_keyword_var.set(config.get("filename_keyword", ""))
        self.attach_excel_var.set(config.get("attach_excel", False))
        
        # 加载自定义占位符
        self.custom_placeholders = config.get("custom_placeholders", {}).copy()
        
        # 如果有Excel文件，加载工作表列表
        base_excel_path = config.get("excel_path", "")
        # 方案一：如果开启智能匹配，尝试查找最新文件
        if self.smart_match_var.get() and base_excel_path:
            latest_path = self.get_latest_file_in_folder(base_excel_path)
            if latest_path and latest_path != base_excel_path:
                print(f"Smart Match: Replaced {base_excel_path} with {latest_path}")
                # Update the variable so user sees the new file
                if self.load_excel_path_var.get():
                     self.excel_path_var.set(latest_path)
                base_excel_path = latest_path
        
        if base_excel_path and os.path.exists(base_excel_path):
            self.load_sheets()
            self.sheet_combo.set(config.get("sheet_name", ""))
        
        self.status_var.set(f"已加载配置: {config_name}")
    
    def delete_config(self):
        """删除选中的配置"""
        config_name = self.config_combo.get()
        if not config_name:
            messagebox.showwarning("警告", "请先选择要删除的配置")
            return
        
        if messagebox.askyesno("确认删除", f"确定要删除配置 '{config_name}' 吗？"):
            del self.configs[config_name]
            if self.save_configs():
                self.update_config_list()
                self.new_config()
                messagebox.showinfo("成功", f"配置 '{config_name}' 已删除")
    
    def select_excel_file(self):
        """选择Excel文件"""
        file_paths = filedialog.askopenfilenames(
            title="选择Excel文件 (支持多选)",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_paths:
            # 将多个路径用分号连接
            paths_str = ";".join(file_paths)
            self.excel_path_var.set(paths_str)
            self.add_to_excel_history(paths_str)
            self.load_sheets()
            if len(file_paths) > 1:
                self.status_var.set(f"已选择 {len(file_paths)} 个文件")
            else:
                self.status_var.set(f"已选择文件: {os.path.basename(file_paths[0])}")
                # 更新默认文件夹
                self.default_excel_folder = os.path.dirname(file_paths[0])
                self.save_ui_preferences()

    def select_excel_folder(self):
        """选择文件夹作为数据源"""
        folder_path = filedialog.askdirectory(title="选择包含Excel文件的文件夹")
        if not folder_path:
            return
            
        try:
            folder = Path(folder_path)
            if not folder.exists():
                return
                
            # 查找Excel文件
            files = []
            keyword = self.filename_keyword_var.get().strip()
            
            # 扩展名列表
            extensions = ['*.xlsx', '*.xls']
            
            for ext in extensions:
                found = list(folder.glob(ext))
                # 关键词过滤
                if keyword:
                    found = [f for f in found if keyword in f.name]
                files.extend(found)
            
            if not files:
                msg = "该文件夹下没有找到"
                if keyword:
                    msg += f"包含关键词 '{keyword}' 的"
                msg += "Excel文件"
                messagebox.showwarning("提示", msg)
                return
            
            # 按修改时间排序，最新的在前
            files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            latest_file = str(files[0])
            
            self.excel_path_var.set(latest_file)
            self.add_to_excel_history(latest_file)
            self.load_sheets()
            
            self.status_var.set(f"已加载文件夹最新文件: {os.path.basename(latest_file)}")
            
            # 更新默认文件夹
            self.default_excel_folder = folder_path
            self.save_ui_preferences()
            
            messagebox.showinfo("成功", f"已选中文件夹并加载最新文件：\n{os.path.basename(latest_file)}\n\n(已自动启用智能文件匹配功能)")
            
            # 自动勾选智能匹配
            self.smart_match_var.set(True)
            
        except Exception as e:
            messagebox.showerror("错误", f"处理文件夹失败: {str(e)}")

    def load_latest_excel_file(self):
        """加载默认文件夹中最新的Excel文件"""
        folder = self.default_excel_folder
        if not folder or not os.path.exists(folder):
            # 如果没有默认文件夹，尝试从历史记录获取
            if self.excel_history:
                folder = os.path.dirname(self.excel_history[0].split(';')[0])
            else:
                # 还是没有，就让用户选一个文件夹
                folder = filedialog.askdirectory(title="选择包含Excel文件的文件夹")
        
        if not folder or not os.path.exists(folder):
            return

        try:
            # 查找所有xlsx和xls文件
            files = []
            for ext in ['*.xlsx', '*.xls']:
                files.extend(list(Path(folder).glob(ext)))
            
            if not files:
                messagebox.showinfo("提示", "该文件夹中没有Excel文件")
                return
            
            # 按修改时间排序，最新的在前
            files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            latest_file = str(files[0])
            
            self.excel_path_var.set(latest_file)
            self.add_to_excel_history(latest_file)
            self.load_sheets()
            self.status_var.set(f"已加载最新文件: {os.path.basename(latest_file)}")
            
            # 更新默认文件夹
            self.default_excel_folder = folder
            self.save_ui_preferences()
            
        except Exception as e:
            messagebox.showerror("错误", f"加载最新文件失败: {str(e)}")

    def smart_search_excel(self):
        """智能查找Excel文件"""
        current_path = self.excel_path_var.get()
        if not current_path:
            messagebox.showwarning("提示", "请先在配置中设置一个Excel文件名，以便系统知道要查找哪个文件")
            return
            
        # 获取文件名
        target_filename = os.path.basename(current_path.split(';')[0])
        
        # 选择搜索目录
        start_dir = self.default_excel_folder if self.default_excel_folder else os.path.expanduser("~")
        folder = filedialog.askdirectory(title=f"选择包含 '{target_filename}' 的文件夹", initialdir=start_dir)
        
        if not folder:
            return
            
        self.status_var.set(f"正在搜索 {target_filename} ...")
        self.root.update()
        
        found_files = []
        try:
            # 递归搜索
            for root, dirs, files in os.walk(folder):
                if target_filename in files:
                    found_files.append(os.path.join(root, target_filename))
            
            if not found_files:
                messagebox.showinfo("搜索结果", f"在选定文件夹中未找到 '{target_filename}'")
                self.status_var.set("未找到文件")
                return
            
            # 如果找到多个，让用户选择（或者默认使用最新的）
            if len(found_files) > 1:
                # 按修改时间排序
                found_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                
            # 使用找到的第一个（最新的）
            new_path = found_files[0]
            self.excel_path_var.set(new_path)
            self.add_to_excel_history(new_path)
            self.load_sheets()
            
            messagebox.showinfo("成功", f"已找到并加载文件：\n{new_path}")
            self.status_var.set(f"已加载: {target_filename}")
            
            # 更新默认文件夹
            self.default_excel_folder = folder
            self.save_ui_preferences()
            
        except Exception as e:
            messagebox.showerror("错误", f"搜索失败: {str(e)}")

    def add_to_excel_history(self, path):
        """添加路径到历史记录"""
        if not path: return
        
        if path in self.excel_history:
            self.excel_history.remove(path)
        
        self.excel_history.insert(0, path)
        # 限制历史记录数量
        if len(self.excel_history) > 10:
            self.excel_history = self.excel_history[:10]
            
        # 更新下拉列表
        if hasattr(self, 'excel_path_combo'):
            self.excel_path_combo['values'] = self.excel_history
        
        self.save_ui_preferences()
    
    def load_sheets(self):
        """加载Excel工作表列表"""
        excel_paths_str = self.excel_path_var.get()
        if not excel_paths_str:
            return

        # 处理路径中的占位符
        excel_paths_str = self.process_template_variables(excel_paths_str)
            
        # 如果有多个文件，只读取第一个文件的Sheet
        excel_path = excel_paths_str.split(';')[0]
        
        if not os.path.exists(excel_path):
            return
        
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            self.sheet_combo['values'] = sheet_names
            if sheet_names:
                self.sheet_combo.current(0)
            wb.close()
        except Exception as e:
            messagebox.showerror("错误", f"读取工作表失败: {str(e)}")
    
    def on_sheet_selected(self, event=None):
        """工作表选择变化时"""
        self.current_excel_data = None
    
    def auto_detect_range(self):
        """自动检测工作表中有数据的范围"""
        excel_paths_str = self.excel_path_var.get()
        # 处理路径中的占位符
        excel_paths_str = self.process_template_variables(excel_paths_str)

        sheet_name = self.sheet_combo.get()
        col_limit_str = self.col_limit_var.get().strip().upper()
        
        if not excel_paths_str:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
            
        # 只检测第一个文件
        excel_path = excel_paths_str.split(';')[0]
        
        if not os.path.exists(excel_path):
            messagebox.showwarning("警告", f"Excel文件不存在: {excel_path}")
            return
        
        if not sheet_name:
            messagebox.showwarning("警告", "请选择工作表")
            return
        
        wb = None
        try:
            self.status_var.set("正在检测数据范围...")
            self.root.update()
            
            # 解析列限制
            target_cols = set()
            if col_limit_str:
                try:
                    from openpyxl.utils import column_index_from_string
                    parts = col_limit_str.replace('，', ',').split(',')
                    for part in parts:
                        part = part.strip()
                        if ':' in part:
                            start, end = part.split(':')
                            start_idx = column_index_from_string(start.strip())
                            end_idx = column_index_from_string(end.strip())
                            for i in range(min(start_idx, end_idx), max(start_idx, end_idx) + 1):
                                target_cols.add(i)
                        else:
                            if part:
                                target_cols.add(column_index_from_string(part))
                except Exception as e:
                    messagebox.showwarning("警告", f"列限制格式错误: {str(e)}")
                    self.status_var.set("就绪")
                    return

            # 重要：不使用 read_only 模式，避免 ReadOnlyWorksheet 的限制
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
            ws = wb[sheet_name]
            
            detected_range = None
            min_row = None
            max_row = None
            min_col = None
            max_col = None
            
            # 使用可靠的方法：直接扫描单元格
            try:
                # 获取工作表的最大范围
                if hasattr(ws, 'max_row') and hasattr(ws, 'max_column'):
                    max_scan_row = ws.max_row if ws.max_row else 1000
                    max_scan_col = ws.max_column if ws.max_column else 100
                else:
                    # 如果无法获取，使用默认范围
                    max_scan_row = 1000
                    max_scan_col = 100
                
                # 限制扫描范围以提高速度
                max_scan_row = min(max_scan_row, 1000)
                max_scan_col = min(max_scan_col, 100)
                
                # 扫描单元格找到有数据的范围
                for row_idx in range(1, max_scan_row + 1):
                    # 确定要扫描的列
                    if target_cols:
                        cols_to_scan = [c for c in target_cols if c <= max_scan_col]
                    else:
                        cols_to_scan = range(1, max_scan_col + 1)
                        
                    for col_idx in cols_to_scan:
                        try:
                            cell = ws.cell(row=row_idx, column=col_idx)
                            # 检查单元格是否有值
                            if cell.value is not None:
                                value_str = str(cell.value).strip()
                                if value_str:  # 不是空字符串
                                    if min_row is None or row_idx < min_row:
                                        min_row = row_idx
                                    if max_row is None or row_idx > max_row:
                                        max_row = row_idx
                                    if min_col is None or col_idx < min_col:
                                        min_col = col_idx
                                    if max_col is None or col_idx > max_col:
                                        max_col = col_idx
                        except Exception:
                            # 跳过无法访问的单元格
                            continue
                
                # 生成范围字符串
                if min_row and max_row and min_col and max_col:
                    from openpyxl.utils import get_column_letter
                    start_cell = f"{get_column_letter(min_col)}{min_row}"
                    end_cell = f"{get_column_letter(max_col)}{max_row}"
                    detected_range = f"{start_cell}:{end_cell}"
                    
            except Exception as e:
                # 如果扫描失败，尝试简单方法
                pass
            
            # 如果检测失败，返回错误
            if not detected_range:
                wb.close()
                msg = "未在工作表中检测到数据"
                if target_cols:
                    msg += f"\n(在指定的列范围内: {col_limit_str})"
                msg += "\n\n可能原因：\n1. 工作表为空\n2. 数据超出扫描范围(1000行)"
                messagebox.showwarning("警告", msg)
                self.status_var.set("未检测到数据")
                return
            
            wb.close()
            
            # 更新范围输入框
            self.range_var.set(detected_range)
            
            # 计算行数和列数
            parts = detected_range.split(':')
            if len(parts) == 2:
                import re
                start_match = re.match(r'([A-Z]+)(\d+)', parts[0])
                end_match = re.match(r'([A-Z]+)(\d+)', parts[1])
                if start_match and end_match:
                    rows = int(end_match.group(2)) - int(start_match.group(2)) + 1
                    
                    # 计算列数
                    def col_to_num(col):
                        num = 0
                        for c in col:
                            num = num * 26 + (ord(c) - ord('A') + 1)
                        return num
                    
                    cols = col_to_num(end_match.group(1)) - col_to_num(start_match.group(1)) + 1
                    
                    messagebox.showinfo("检测成功", 
                                       f"检测到数据范围: {detected_range}\n\n"
                                       f"行数: {rows}\n"
                                       f"列数: {cols}\n\n"
                                       f"点击'读取数据'以加载数据")
            else:
                messagebox.showinfo("检测成功", f"检测到数据范围: {detected_range}")
            
            self.status_var.set(f"已检测到范围: {detected_range}")
            
        except Exception as e:
            # 确保关闭工作簿
            if wb:
                try:
                    wb.close()
                except:
                    pass
            
            import traceback
            error_details = traceback.format_exc()
            messagebox.showerror("错误", f"自动检测范围失败:\n{str(e)}\n\n详细信息:\n{error_details}")
            self.status_var.set("检测失败")
        finally:
            # 最终确保工作簿被关闭
            if wb:
                try:
                    wb.close()
                except:
                    pass
    
    def _internal_read_excel(self, excel_path, sheet_name, data_range):
        """内部方法：读取指定Excel文件的内容，返回数据列表"""
        if not os.path.exists(excel_path):
             return None, f"文件不存在: {excel_path}"
             
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return None, f"工作表 {sheet_name} 不存在"
                
            ws = wb[sheet_name]
            
            # 读取指定范围的数据
            try:
                cells = ws[data_range]
            except Exception as e:
                wb.close()
                return None, f"无法读取范围 {data_range}: {e}"
                
            current_file_rows = []
            
            # 处理单行或多行数据 (openpyxl range handling)
            if isinstance(cells, tuple) and len(cells) > 0:
                if isinstance(cells[0], tuple):
                    # 多行数据
                    for row in cells:
                        row_data = [cell.value if cell.value is not None else "" for cell in row]
                        current_file_rows.append(row_data)
                else:
                    # 单行数据
                    row_data = [cell.value if cell.value is not None else "" for cell in cells]
                    current_file_rows.append(row_data)
            elif hasattr(cells, 'value'): # 单个单元格
                 current_file_rows.append([cells.value if cells.value is not None else ""])
            else:
                 # 可能是 list of cells
                 if isinstance(cells, list):
                     row_data = [cell.value if cell.value is not None else "" for cell in cells]
                     current_file_rows.append(row_data)

            # 尝试获取列宽 (仅作参考)
            column_widths = []
            try:
                # 解析起始列（如 A1:F10 中的 A）
                if ':' in data_range:
                    start_col_str = data_range.split(':')[0]
                else:
                    start_col_str = data_range
                    
                col_letter_start = ''.join([c for c in start_col_str if c.isalpha()])
                from openpyxl.utils import column_index_from_string, get_column_letter
                if col_letter_start:
                    start_col_idx = column_index_from_string(col_letter_start)
                    
                    if current_file_rows and len(current_file_rows) > 0:
                         # 读取每列的实际宽度
                        for i in range(len(current_file_rows[0])):
                            col_idx = start_col_idx + i
                            col_letter_current = get_column_letter(col_idx)
                            if col_letter_current in ws.column_dimensions:
                                width = ws.column_dimensions[col_letter_current].width
                                # Excel宽度转像素 (近似值)
                                pixel_width = int(width * 7) if width else 80
                                column_widths.append(pixel_width)
                            else:
                                column_widths.append(80)
            except:
                pass

            wb.close()
            return (current_file_rows, column_widths), None

        except Exception as e:
            return None, str(e)

    def read_excel_data(self):
        """读取Excel数据"""
        excel_paths_str = self.excel_path_var.get()
        # 处理路径中的占位符
        excel_paths_str = self.process_template_variables(excel_paths_str)

        sheet_name = self.sheet_combo.get()
        data_range = self.range_var.get().strip()
        
        if not excel_paths_str:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        if not sheet_name:
            messagebox.showwarning("警告", "请选择工作表")
            return
        
        if not data_range:
            messagebox.showwarning("警告", "请输入数据范围")
            return
            
        file_paths = excel_paths_str.split(';')
        all_data_rows = []
        headers = None
        
        self.status_var.set("正在读取数据...")
        self.root.update()
        
        try:
            for idx, excel_path in enumerate(file_paths):
                # 调用内部方法
                result, error = self._internal_read_excel(excel_path, sheet_name, data_range)
                
                if error:
                    print(f"Warning reading {excel_path}: {error}")
                    continue
                
                if not result:
                    continue
                    
                current_file_rows, col_widths = result
                
                # 记录第一个文件的列宽
                if idx == 0:
                    self.excel_column_widths = col_widths
                
                if not current_file_rows:
                    continue

                # 合并逻辑：如果是第一个文件，保留表头；后续文件跳过表头
                if idx == 0:
                    all_data_rows.extend(current_file_rows)
                    if current_file_rows:
                        headers = current_file_rows[0]
                else:
                    # 检查表头是否一致（可选）
                    if headers and len(current_file_rows) > 0 and current_file_rows[0] == headers:
                        # 简单的合并：跳过第一行（假设是表头）
                        all_data_rows.extend(current_file_rows[1:])
                    else:
                        all_data_rows.extend(current_file_rows)
        
            if not all_data_rows:
                messagebox.showinfo("提示", "未读取到数据或文件不存在")
                self.current_excel_data = None
            else:
                self.current_excel_data = all_data_rows
                self.status_var.set(f"读取成功: {len(all_data_rows)} 行数据 (来自 {len(file_paths)} 文件)")
                messagebox.showinfo("成功", f"成功读取 {len(all_data_rows)} 行数据\n(包含表头)")
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("错误", f"读取Excel失败: {str(e)}")
            self.status_var.set("读取失败")

    # 移除 open_current_folder 方法，因为它已不再被使用
    
    def format_excel_data_as_table(self):
        """将Excel数据格式化为HTML表格"""
        if not self.current_excel_data:
            return "[未读取Excel数据]"
        
        html = "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse;'>\n"
        
        for i, row in enumerate(self.current_excel_data):
            html += "  <tr>\n"
            for cell in row:
                cell_value = str(cell) if cell else ""
                if i == 0:  # 第一行作为表头
                    html += f"    <th style='background-color: #4472C4; color: white; font-weight: bold;'>{cell_value}</th>\n"
                else:
                    html += f"    <td>{cell_value}</td>\n"
            html += "  </tr>\n"
        
        html += "</table>"
        return html
    
    def format_excel_data_as_text(self):
        """将Excel数据格式化为纯文本"""
        if not self.current_excel_data:
            return "[未读取Excel数据]"
        
        text_lines = []
        for row in self.current_excel_data:
            row_text = "\t".join([str(cell) if cell else "" for cell in row])
            text_lines.append(row_text)
        
        return "\n".join(text_lines)
    
    def get_latest_file_in_folder(self, excel_path, keyword=None):
        """
        方案一（增强版）：
        1. 关键词寻址：如果设置了关键词，只匹配包含关键词的文件。
        2. 占位符处理：如果文件名包含未解析的 {xxx}，自动转换为通配符 * 进行搜索。
        3. 智能文件夹切换：检测是否有更新的同级文件夹（如 2024-01 -> 2024-02），如果有则切换。
        Params:
            keyword (str, optional): 覆盖UI的关键词设置，用于批量处理时传入特定配置的关键词。
        """
        if not excel_path:
            return None
        
        try:
            path_obj = Path(excel_path)
            
            # --- 增强逻辑：处理路径中未解析的占位符 ---
            # 如果路径本身不存在，可能是因为包含占位符。我们尝试解析它。
            # 如果 folder 部分包含占位符且不存在，我们无能为力（除非遍历上级目录，太复杂暂不处理）
            # 我们重点处理 filename 部分的占位符
            
            current_folder = path_obj.parent
            if not current_folder.exists():
                return excel_path # 文件夹都不存在，直接返回原路径（后续会报错）

            suffix = path_obj.suffix
            stem = path_obj.stem
            
            # 自动生成通配符模式：将 MyFile_{data}.xlsx 转换为 MyFile_*.xlsx
            # 只有当文件名包含 { } 时才触发
            wildcard_pattern = ""
            if "{" in stem and "}" in stem:
                wildcard_pattern = re.sub(r'\{.*?\}', '*', stem) + suffix
                # print(f"DEBUG: Converted placeholder path to wildcard: {wildcard_pattern}")
            
            # 获取配置的关键词 (优先使用传入参数，否则使用UI)
            if keyword is None:
                keyword = self.filename_keyword_var.get().strip()
            else:
                keyword = keyword.strip()
            
            # --- 步骤 1: 优先在当前定位文件夹中查找 ---
            
            # 策略 A: 精确匹配 (带关键词)
            if keyword:
                search_pattern = f"*{keyword}*{suffix}"
                files = list(current_folder.glob(search_pattern))
                if files:
                    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                    return str(files[0])
            
            # 策略 B: 使用文件名转换的通配符 (处理 {data} -> *)
            if wildcard_pattern:
                files = list(current_folder.glob(wildcard_pattern))
                if files:
                    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                    # print(f"DEBUG: Found file using wildcard: {files[0]}")
                    return str(files[0])
            
            # 策略 C: 如果没找到带关键词的，或者没设置关键词 -> 查找最新的文件
            # 用户要求："如果没有,再尋找文件里面最新修改时间的文件"
            fallback_pattern = f"*{suffix}"
            files = list(current_folder.glob(fallback_pattern))
            if files:
                files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                return str(files[0])

            # --- 步骤 2: (备选) 如果当前文件夹没找到任何Excel，尝试智能跨文件夹搜索 ---
            # 检查是否有更新的同级文件夹
            try:
                parent_dir = current_folder.parent
                if parent_dir.exists():
                    subdirs = [d for d in parent_dir.iterdir() if d.is_dir()]
                    subdirs = [d for d in subdirs if not d.name.startswith('.')]
                    
                    if subdirs:
                        subdirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                        newest_dir = subdirs[0]
                        
                        if newest_dir != current_folder:
                            # 在最新文件夹里找
                            # 2.1 关键词
                            if keyword:
                                files = list(newest_dir.glob(f"*{keyword}*{suffix}"))
                                if files:
                                    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                                    return str(files[0])
                            
                            # 2.2 最新文件
                            files = list(newest_dir.glob(fallback_pattern))
                            if files:
                                files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                                return str(files[0])
            except Exception as e:
                print(f"Folder search error: {e}")
            
            return excel_path
            
        except Exception as e:
            print(f"Smart search error: {e}")
            return excel_path

    def process_smart_image_logic(self, body_template, excel_path, sheet_name, data_range):
        """
        方案二：处理 {SMART_IMAGE} 占位符
        返回: (processed_body, new_inline_image_dict or None)
        """
        if "{SMART_IMAGE}" not in body_template:
            return body_template, None
            
        if not excel_path or not os.path.exists(excel_path) or not sheet_name or not data_range:
            return body_template.replace("{SMART_IMAGE}", "[错误: 无法获取Excel截图，请检查配置]"), None
            
        # 截图
        img_path = self.copy_range_as_image_com(excel_path, sheet_name, data_range)
        if not img_path:
             return body_template.replace("{SMART_IMAGE}", "[错误: 截图失败]"), None
             
        # 生成CID
        cid = f"smart_img_{uuid.uuid4().hex}@outlook.drafter"
        
        # 替换正文
        new_body = body_template.replace("{SMART_IMAGE}", f'<img src="cid:{cid}" style="max-width:100%">')
        
        return new_body, {'path': img_path, 'cid': cid}

    def generate_preview(self):
        """在弹出窗口中生成邮件预览"""
        # 获取邮箱地址并处理变量
        to_raw = self.to_entry_var.get()
        cc_raw = self.cc_entry_var.get()
        bcc_raw = self.bcc_entry_var.get()
        
        to_processed = self.process_template_variables(to_raw)
        cc_processed = self.process_template_variables(cc_raw)
        bcc_processed = self.process_template_variables(bcc_raw)
        
        to_list = [x.strip() for x in to_processed.split(';') if x.strip()]
        cc_list = [x.strip() for x in cc_processed.split(';') if x.strip()]
        bcc_list = [x.strip() for x in bcc_processed.split(';') if x.strip()]
        
        subject = self.process_template_variables(self.subject_var.get().strip())
        body_template = self.body_text.get(1.0, tk.END).strip()
        
        if not to_list:
            messagebox.showwarning("警告", "请添加至少一个收件人")
            return
        
        if not subject:
            messagebox.showwarning("警告", "请填写主题")
            return
        
        # 创建预览窗口
        preview_window = tk.Toplevel(self.root)
        preview_window.title("邮件预览")
        preview_window.geometry("800x600")
        preview_window.transient(self.root)
        
        # 标题栏
        title_frame = ttk.Frame(preview_window, padding="10")
        title_frame.pack(fill=tk.X)
        ttk.Label(title_frame, text="📧 邮件预览", font=('TkDefaultFont', 12, 'bold')).pack(anchor=tk.W)
        
        # 预览内容
        preview_frame = ttk.Frame(preview_window, padding="10")
        preview_frame.pack(fill=tk.BOTH, expand=True)
        
        # 替换Excel数据占位符
        if "{EXCEL_DATA}" in body_template:
            excel_html = self.format_excel_data_as_table()
            body = body_template.replace("{EXCEL_DATA}", excel_html)
        else:
            body = body_template
        
        # 生成预览文本
        preview = f"{'='*60}\n"
        preview += f"收件人: {'; '.join(to_list)}\n"
        if cc_list:
            preview += f"抄送: {'; '.join(cc_list)}\n"
        if bcc_list:
            preview += f"密送: {'; '.join(bcc_list)}\n"
        preview += f"主题: {subject}\n"
        preview += f"{'='*60}\n\n"
        preview += f"正文预览:\n{'-'*60}\n"
        
        # 显示纯文本版本的预览
        if "{EXCEL_DATA}" in body_template and self.current_excel_data:
            preview_body = self.process_template_variables(body_template)
            preview += preview_body.replace("{EXCEL_DATA}", "\n" + self.format_excel_data_as_text() + "\n")
        else:
            preview += self.process_template_variables(body)
        
        preview += f"\n{'-'*60}\n"
        if self.attachments:
            preview += f"\n附件: {len(self.attachments)} 个文件\n"
            for att in self.attachments[:5]:
                preview += f"  • {os.path.basename(att)}\n"
            if len(self.attachments) > 5:
                preview += f"  ... 还有 {len(self.attachments)-5} 个附件\n"
        preview += f"\n注意: 实际邮件将使用HTML格式显示表格\n"
        
        # 预览文本框
        preview_text = scrolledtext.ScrolledText(preview_frame, width=80, height=30, wrap=tk.WORD, font=('Consolas', 9))
        preview_text.pack(fill=tk.BOTH, expand=True)
        preview_text.insert(1.0, preview)
        preview_text.config(state='disabled')
        
        # 按钮区
        button_frame = ttk.Frame(preview_window, padding="10")
        button_frame.pack(fill=tk.X)
        ttk.Button(button_frame, text="关闭", command=preview_window.destroy, width=15).pack(side=tk.RIGHT, padx=5)
        
        self.status_var.set("预览已生成")
    
    def create_draft(self):
        """创建Outlook草稿"""
        # 获取并处理邮箱地址中的变量
        to_raw = self.to_entry_var.get()
        cc_raw = self.cc_entry_var.get()
        bcc_raw = self.bcc_entry_var.get()
        
        to_processed = self.process_template_variables(to_raw)
        cc_processed = self.process_template_variables(cc_raw)
        bcc_processed = self.process_template_variables(bcc_raw)
        
        # 用于验证
        to_list = [x.strip() for x in to_processed.split(';') if x.strip()]
        
        subject = self.subject_var.get().strip()
        body_template = self.body_text.get(1.0, tk.END).strip()
        
        if not to_list:
            messagebox.showwarning("警告", "请添加至少一个收件人")
            return
        
        if not subject:
            messagebox.showwarning("警告", "请填写主题")
            return
        
        # 方案一：智能匹配最新文件
        processed_excel_paths = self.process_template_variables(self.excel_path_var.get())
        current_excel_path = processed_excel_paths.split(';')[0]
        sheet_name = self.sheet_combo.get()
        data_range = self.range_var.get()
        
        if self.smart_match_var.get() and current_excel_path:
            latest_path = self.get_latest_file_in_folder(current_excel_path)
            if latest_path != current_excel_path:
                print(f"Creating Draft: Switching {current_excel_path} -> {latest_path}")
                current_excel_path = latest_path

        # 方案二：处理 {SMART_IMAGE}
        # 注意：这会生成新的临时文件，不应保存到 self.inline_images 以免污染后续操作，
        # 而是应该只为此单次创建添加。
        temp_inline_images = []
        if "{SMART_IMAGE}" in body_template:
            body_template, smart_img_info = self.process_smart_image_logic(
                body_template, current_excel_path, sheet_name, data_range
            )
            if smart_img_info:
                temp_inline_images.append(smart_img_info)

        try:
            # 连接Outlook
            outlook = win32com.client.Dispatch("Outlook.Application")
            
            # 创建邮件对象
            mail = outlook.CreateItem(0)  # 0 代表邮件项
            
            # 设置收件人
            mail.To = to_processed
            if cc_processed:
                mail.CC = cc_processed
            if bcc_processed:
                mail.BCC = bcc_processed
            
            # 设置优先级和已读回执
            if self.priority_var.get():
                mail.Importance = 2  # 2 = olImportanceHigh
            
            if self.receipt_var.get():
                mail.ReadReceiptRequested = True
            
            # 设置主题
            subject_processed = self.process_template_variables(subject)
            mail.Subject = subject_processed
            
            # 设置正文（HTML格式）
            body_html = self.process_template_variables(body_template)
            if "{EXCEL_DATA}" in body_html:
                excel_html = self.format_excel_data_as_table()
                body_html = body_html.replace("{EXCEL_DATA}", excel_html)
            
            # 处理 {EXCEL_IMAGE} 占位符
            if "{EXCEL_IMAGE}" in body_html:
                try:
                    generated_images = self.generate_excel_images()
                    if generated_images:
                        img_fragments = []
                        for img_path in generated_images:
                            cid = f"img_{uuid.uuid4().hex}@local"
                            self.inline_images.append({'path': img_path, 'cid': cid})
                            img_tag = f'<img src="cid:{cid}" alt="Excel表格" style="max-width:100%; border:1px solid #ccc;"><br/>'
                            img_fragments.append(img_tag)
                        
                        body_html = body_html.replace("{EXCEL_IMAGE}", "\n".join(img_fragments))
                except Exception as e:
                    print(f"生成图片失败: {e}")

            # 将纯文本转换为HTML（保留换行）
            body_html = body_html.replace("\n", "<br>")
            mail.HTMLBody = body_html
            
            # 添加普通附件（跳过已作为 inline_images 的路径，避免重复）
            inline_paths = [img.get('path') for img in getattr(self, 'inline_images', [])]
            
            # 如果勾选了"附带源文件"，则将当前使用的Excel文件加入附件列表
            final_attachments = self.attachments.copy()
            if self.attach_excel_var.get() and current_excel_path and os.path.exists(current_excel_path):
                # 避免重复添加
                if current_excel_path not in final_attachments:
                    final_attachments.append(current_excel_path)

            for attachment_path in final_attachments:
                if attachment_path in inline_paths:
                    continue
                if os.path.exists(attachment_path):
                    mail.Attachments.Add(attachment_path)

            # 添加并标记内嵌图片（Content-ID），以便在HTML中使用cid:引用
            all_inline_images = getattr(self, 'inline_images', []) + temp_inline_images
            for img in all_inline_images:
                path = img.get('path')
                cid = img.get('cid')
                if not path or not cid:
                    continue
                if os.path.exists(path):
                    try:
                        att = mail.Attachments.Add(path)
                        # PR_ATTACH_CONTENT_ID (0x3712001F) 指定附件的 Content-ID
                        try:
                            # 关键：设置为隐藏附件 (0x7FFE000B = PR_ATTACHMENT_HIDDEN)
                            # 这有助于实现“粘贴图片不带到附件栏”的效果
                            # 但注意：Outlook行为不一，通常 CID + HTML引用 就足够隐藏
                            # att.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x7FFE000B", True)
                            
                            att.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)
                        except Exception:
                            # 忽略无法设置属性的情况
                            pass
                    except Exception:
                        pass
            
            # 保存为草稿（不发送）
            mail.Save()
            mail.Display() # 显示草稿窗口

            # 清理临时生成的智能截图
            for img in temp_inline_images:
                try:
                    if os.path.exists(img['path']):
                        os.remove(img['path'])
                except:
                    pass
            
            # 清理内嵌图片记录（防止传递到下一个草稿）
            self.cleanup_inline_images()
            
            # 记录历史
            self.save_to_history("; ".join(to_list), subject_processed)
            
            messagebox.showinfo("成功", "草稿邮件已创建！\n请在Outlook草稿箱中查看。")
            self.status_var.set(f"草稿已创建 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
        except Exception as e:
            messagebox.showerror("错误", f"创建草稿失败: {str(e)}\n\n请确保已安装并登录Outlook。")
    
    def add_attachment(self):
        """添加附件"""
        files = filedialog.askopenfilenames(
            title="选择附件",
            filetypes=[("所有文件", "*.*")]
        )
        for file_path in files:
            if file_path and file_path not in self.attachments:
                self.attachments.append(file_path)
        self.update_attachment_list()
        self.status_var.set(f"已添加 {len(files)} 个附件")

    def attach_current_excel(self):
        """将当前选中的 Excel 文件添加为附件"""
        excel_paths_str = self.excel_path_var.get()
        if not excel_paths_str:
             messagebox.showwarning("警告", "请先选择 Excel 文件")
             return
        
        # 处理路径中的占位符
        excel_paths_str = self.process_template_variables(excel_paths_str)

        file_paths = excel_paths_str.split(';')
        count = 0
        for path in file_paths:
            path = path.strip()
            if path and os.path.exists(path):
                if path not in self.attachments:
                    self.attachments.append(path)
                    count += 1
        
        if count > 0:
            self.update_attachment_list()
            self.status_var.set(f"已添加 {count} 个 Excel 文件作为附件")
        else:
            messagebox.showinfo("提示", "Excel 文件已在附件列表中")
    
    def clear_attachments(self):
        """清空附件列表"""
        self.attachments = []
        self.update_attachment_list()
        self.status_var.set("已清空附件")
    
    def update_attachment_list(self):
        """更新附件列表显示"""
        self.attachment_listbox.delete(0, tk.END)
        for attachment in self.attachments:
            filename = os.path.basename(attachment)
            # 标注是否为内嵌图片
            is_inline = any(img.get('path') == attachment for img in getattr(self, 'inline_images', []))
            display_name = filename + ("  [内嵌]" if is_inline else "")
            self.attachment_listbox.insert(tk.END, display_name)
    
    def cleanup_inline_images(self):
        """清理内嵌图片及其临时文件"""
        if not hasattr(self, 'inline_images'):
            return
        
        # 删除临时图片文件
        for img in self.inline_images:
            path = img.get('path')
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass  # 忽略删除失败
            
            # 从附件列表中移除
            if path in self.attachments:
                self.attachments.remove(path)
        
        # 清空inline_images列表
        self.inline_images = []
        
        # 更新附件显示
        self.update_attachment_list()

    def paste_with_formatting(self):
        """保留源格式粘贴 Excel 数据到正文（Keep Source Formatting）"""
        if not self.current_excel_data:
            messagebox.showwarning("警告", "请先读取 Excel 数据")
            return
        
        if len(self.current_excel_data) == 0:
            messagebox.showwarning("警告", "Excel 数据为空")
            return
        
        try:
            self.status_var.set("正在生成格式化表格...")
            self.root.update()
            
            # 生成带格式的 HTML 表格（类似 Excel 的样式）
            html_table = self.generate_formatted_html_table(self.current_excel_data)
            
            # 获取当前正文内容
            try:
                body = self.body_text.get(1.0, tk.END).strip()
            except:
                body = ""
            
            # 如果正文存在 {EXCEL_DATA} 占位符，替换它；否则追加到结尾
            if "{EXCEL_DATA}" in body:
                new_body = body.replace("{EXCEL_DATA}", html_table)
                self.body_text.delete(1.0, tk.END)
                self.body_text.insert(1.0, new_body)
            else:
                # 追加到结尾
                if body:
                    self.body_text.insert(tk.END, "\n\n" + html_table)
                else:
                    self.body_text.insert(1.0, html_table)
            
            messagebox.showinfo("成功", f"已插入格式化表格（{len(self.current_excel_data)}行）\n\n使用了 Keep Source Formatting 样式")
            self.status_var.set("已插入格式化表格")
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            messagebox.showerror("错误", f"插入表格失败:\n{str(e)}\n\n详细信息:\n{error_details}")
            self.status_var.set("插入失败")
    
    def copy_range_as_image_com(self, excel_path, sheet_name, data_range):
        """使用 COM 接口调用 Excel 复制范围为图片 (保留原格式)"""
        if ImageGrab is None:
            return None

        excel = None
        wb = None
        try:
            # 初始化 Excel 应用
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
            except:
                excel = win32com.client.Dispatch("Excel.Application")
            
            # 打开工作簿
            abs_path = os.path.abspath(excel_path)
            wb_opened = False
            # 检查是否已经打开
            try:
                for w in excel.Workbooks:
                    if w.FullName.lower() == abs_path.lower():
                        wb = w
                        wb_opened = True
                        break
            except:
                pass
            
            if not wb:
                wb = excel.Workbooks.Open(abs_path, ReadOnly=True)
            
            # 获取工作表和范围
            try:
                ws = wb.Sheets(sheet_name)
                rng = ws.Range(data_range)
                
                # 复制为图片 (Appearance: 1=xlScreen, Format: 2=xlBitmap)
                rng.CopyPicture(Appearance=1, Format=2)
                
                # 等待剪贴板更新
                time.sleep(0.5)
                
                # 从剪贴板获取图片
                img = ImageGrab.grabclipboard()
                
                if img:
                    # 保存到临时文件
                    temp_dir = tempfile.gettempdir()
                    img_filename = f"excel_paste_{uuid.uuid4().hex}.png"
                    img_path = os.path.join(temp_dir, img_filename)
                    img.save(img_path, 'PNG')
                    return img_path
            finally:
                # 如果是我们打开的，关闭它
                if wb and not wb_opened:
                    wb.Close(SaveChanges=False)
            
            return None
                
        except Exception as e:
            print(f"COM Error: {e}")
            return None

    def generate_excel_images(self):
        """生成Excel截图，返回图片路径列表"""
        excel_paths_str = self.excel_path_var.get()
        # 处理路径中的占位符
        excel_paths_str = self.process_template_variables(excel_paths_str)

        sheet_name = self.sheet_combo.get()
        data_range = self.range_var.get().strip()
        
        if not excel_paths_str or not sheet_name or not data_range:
             return []

        file_paths = excel_paths_str.split(';')
        generated_images = []
        
        for excel_path in file_paths:
            if not os.path.exists(excel_path):
                continue
            
            # 尝试使用 COM 获取精确截图
            img_path = self.copy_range_as_image_com(excel_path, sheet_name, data_range)
            
            # 如果 COM 失败，回退到手动绘制
            if not img_path and self.current_excel_data:
                    img_path = self.create_excel_image(self.current_excel_data)
            
            if img_path and os.path.exists(img_path):
                generated_images.append(img_path)
        
        if not generated_images and self.current_excel_data:
            # 如果没有生成图片，尝试只用 current_excel_data 生成
            img_path = self.create_excel_image(self.current_excel_data)
            if img_path:
                generated_images.append(img_path)
                
        return generated_images

    def paste_as_picture(self):
        """插入 {SMART_IMAGE} 占位符"""
        # 获取当前设置
        if not self.excel_path_var.get() or not self.sheet_combo.get() or not self.range_var.get():
             messagebox.showwarning("警告", "请先选择 Excel 文件、工作表和数据范围")
             return

        try:
            # 简化逻辑：直接插入 {SMART_IMAGE} 占位符
            self.status_var.set("已插入智能截图占位符")
            
            # 获取当前正文内容
            try:
                body = self.body_text.get(1.0, tk.END).strip()
            except:
                body = ""
            
            placeholder = "{SMART_IMAGE}"
            
            # 使用不区分大小写的检查
            if placeholder.lower() in body.lower():
                messagebox.showinfo("提示", f"{placeholder} 占位符已存在于正文中。")
                return

            # 如果存在 {EXCEL_DATA} 或 {EXCEL_IMAGE}，询问替换
            if "{EXCEL_DATA}" in body:
                new_body = body.replace("{EXCEL_DATA}", placeholder)
                self.body_text.delete(1.0, tk.END)
                self.body_text.insert(1.0, new_body)
            elif "{EXCEL_IMAGE}" in body:
                new_body = body.replace("{EXCEL_IMAGE}", placeholder)
                self.body_text.delete(1.0, tk.END)
                self.body_text.insert(1.0, new_body)
            else:
                # 追加到光标位置或结尾
                self.body_text.insert(tk.INSERT, "\n" + placeholder + "\n")
            
            messagebox.showinfo("成功", f"已插入 {placeholder} 占位符。\n\n在生成草稿时，此处将自动替换为最新的 Excel 截图。")
            
        except Exception as e:
            messagebox.showerror("错误", f"操作失败: {str(e)}")
    
    def generate_formatted_html_table(self, data_rows):
        """生成带格式的 HTML 表格（模拟 Keep Source Formatting）- 使用 Excel 实际列宽"""
        html = '<table border="1" cellpadding="8" cellspacing="0" style="border-collapse:collapse; font-family:Calibri,Arial,sans-serif; font-size:11pt;">\n'
        
        # 如果有列宽信息，添加 colgroup
        if hasattr(self, 'excel_column_widths') and self.excel_column_widths:
            html += '  <colgroup>\n'
            for width in self.excel_column_widths:
                html += f'    <col style="width:{width}px;">\n'
            html += '  </colgroup>\n'
        
        for i, row in enumerate(data_rows):
            html += '  <tr>\n'
            for j, cell in enumerate(row):
                cell_value = str(cell) if cell else ""
                # 获取列宽（如果有）
                width_style = ""
                if hasattr(self, 'excel_column_widths') and self.excel_column_widths and j < len(self.excel_column_widths):
                    width_style = f" width:{self.excel_column_widths[j]}px;"
                
                if i == 0:  # 表头
                    html += f'    <th style="background-color:#4472C4; color:white; font-weight:bold; text-align:left; padding:8px; border:1px solid #2E5C8A;{width_style}">{cell_value}</th>\n'
                else:  # 数据行
                    html += f'    <td style="background-color:white; color:#000000; padding:8px; border:1px solid #D0D0D0;{width_style}">{cell_value}</td>\n'
            html += '  </tr>\n'
        
        html += '</table>'
        return html
    
    def create_excel_image(self, data_rows):
        """创建 Excel 表格图片"""
        try:
            if Image is None or ImageDraw is None or ImageFont is None:
                raise RuntimeError("Pillow 未安装或导入失败。\n请运行: pip install pillow")
            
            if not data_rows or len(data_rows) == 0:
                raise ValueError("数据为空")
            
            # 计算列宽
            cols = max((len(r) for r in data_rows), default=0)
            if cols == 0:
                raise ValueError("数据列数为0")
            
            # 优先使用 Excel 实际列宽，否则根据内容计算
            if hasattr(self, 'excel_column_widths') and self.excel_column_widths and len(self.excel_column_widths) >= cols:
                # 使用 Excel 实际列宽
                col_pixel_widths = self.excel_column_widths[:cols]
            else:
                # 根据内容长度计算列宽
                col_widths = [0] * cols
                for row in data_rows:
                    for i in range(min(cols, len(row))):
                        cell = str(row[i]) if row[i] is not None else ""
                        col_widths[i] = max(col_widths[i], len(cell))
                
                # 后续会用到 col_pixel_widths
                char_width_temp = 8  # 临时值，后面会重新计算
                col_pixel_widths = [max(80, int(w * char_width_temp + 20)) for w in col_widths]
            
            # 加载字体
            font_size = 14
            font = None
            font_paths = [
                "C:/Windows/Fonts/msyh.ttc",
                "C:/Windows/Fonts/simsun.ttc",
                "C:/Windows/Fonts/Arial.ttf",
            ]
            for fp in font_paths:
                try:
                    if os.path.exists(fp):
                        font = ImageFont.truetype(fp, font_size)
                        break
                except Exception:
                    continue
            
            # 如果没有找到字体，使用默认
            if font is None:
                try:
                    font = ImageFont.load_default()
                except Exception:
                    font = None
            
            # 计算图片尺寸
            padding = 10
            # 如果没有使用 Excel 列宽，则根据字体大小重新计算
            if not (hasattr(self, 'excel_column_widths') and self.excel_column_widths and len(self.excel_column_widths) >= cols):
                char_width = font_size * 0.6 if font_size else 8
                col_pixel_widths = [max(80, int(w * char_width + 20)) for w in col_widths]
            
            table_width = sum(col_pixel_widths) + padding * 2
            row_height = font_size + 12
            table_height = row_height * len(data_rows) + padding * 2
            
            # 创建图片
            img = Image.new('RGB', (table_width, table_height), color='white')
            draw = ImageDraw.Draw(img)
            
            # 绘制表格
            x = padding
            y = padding
            for ri, row in enumerate(data_rows):
                x = padding
                for ci in range(cols):
                    cell = str(row[ci]) if ci < len(row) and row[ci] is not None else ""
                    w = col_pixel_widths[ci]
                    
                    try:
                        if ri == 0:  # 表头
                            # 绘制背景
                            draw.rectangle([x, y, x + w, y + row_height], fill='#4472C4', outline='#2E5C8A')
                            # 绘制额外边框使其更粗（模拟 width=2）
                            draw.rectangle([x+1, y+1, x + w-1, y + row_height-1], outline='#2E5C8A')
                            # 绘制文字
                            draw.text((x + 6, y + 4), cell, fill='white', font=font)
                        else:  # 数据行
                            draw.rectangle([x, y, x + w, y + row_height], fill='white', outline='#D0D0D0')
                            draw.text((x + 6, y + 4), cell, fill='black', font=font)
                    except Exception as e:
                        # 如果绘制失败，跳过该单元格
                        pass
                    
                    x += w
                y += row_height
            
            # 保存到临时文件
            tmp_dir = tempfile.gettempdir()
            fname = f"excel_picture_{uuid.uuid4().hex}.png"
            path = os.path.join(tmp_dir, fname)
            img.save(path, 'PNG')
            return path
            
        except Exception as e:
            # 捕获所有异常并重新抛出，带详细信息
            raise RuntimeError(f"创建图片失败: {str(e)}")

    def remove_selected_attachment(self, event=None):
        """删除列表中选中的附件（可多选）"""
        selection = self.attachment_listbox.curselection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要删除的附件")
            return
        # 从后往前删除以避免索引漂移
        for index in reversed(selection):
            try:
                del self.attachments[index]
            except Exception:
                pass
        self.update_attachment_list()
        self.status_var.set("已删除选中附件")

    def open_selected_attachment(self, event=None):
        """在系统中打开双击的附件（Windows）"""
        try:
            selection = self.attachment_listbox.curselection()
            if not selection:
                return
            index = selection[0]
            path = self.attachments[index]
            if os.path.exists(path):
                # 使用os.startfile 在 Windows 上打开
                try:
                    os.startfile(path)
                except Exception as e:
                    messagebox.showerror("错误", f"打开附件失败: {e}")
            else:
                messagebox.showwarning("警告", "文件不存在: " + path)
        except Exception as e:
            messagebox.showerror("错误", f"操作失败: {e}")
    
    def process_template_variables(self, text, row_data=None, headers=None):
        """处理模板变量 (增强版：不区分大小写)
           row_data: 指定覆盖的Excel行数据 (用于批量生成)
           headers: 指定表头 (用于批量生成)
        """
        if not text:
            return ""
            
        # 预处理：大小写不敏感替换内置变量
        # 构建映射避免多次遍历
        replacements = {
            "{date}": datetime.now().strftime("%Y-%m-%d"),
            "{time}": datetime.now().strftime("%H:%M:%S"),
            "{datetime}": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # 使用正则进行不区分大小写的替换
        for key, val in replacements.items():
            pattern = re.compile(re.escape(key), re.IGNORECASE)
            text = pattern.sub(val, text)
        
        # 替换自定义占位符
        for name, value in self.custom_placeholders.items():
            placeholder = "{" + name + "}"
            # 尝试不区分大小写匹配自定义占位符名
            pattern = re.compile(re.escape(placeholder), re.IGNORECASE)
            text = pattern.sub(str(value), text)
        
        # 替换签名作为占位符
        if hasattr(self, 'email_signatures'):
            for name, value in self.email_signatures.items():
                placeholder = "{" + name + "}"
                pattern = re.compile(re.escape(placeholder), re.IGNORECASE)
                text = pattern.sub(str(value), text)
        
        # 处理Excel数据
        # 1. 优先使用传入的行数据 (Batch 模式)
        if row_data is not None and headers is not None:
             for i, header in enumerate(headers):
                if i < len(row_data):
                    placeholder = "{" + str(header) + "}"
                    # 不区分大小写的列名匹配
                    pattern = re.compile(re.escape(placeholder), re.IGNORECASE)
                    text = pattern.sub(str(row_data[i]), text)
                    
        # 2. 否则使用当前加载的第一行数据 (Single Draft 模式)
        elif self.current_excel_data and len(self.current_excel_data) > 0:
            current_headers = self.current_excel_data[0] if len(self.current_excel_data) > 0 else []
            if len(self.current_excel_data) > 1:
                first_data_row = self.current_excel_data[1]
                for i, header in enumerate(current_headers):
                    if i < len(first_data_row):
                        placeholder = "{" + str(header) + "}"
                        # 统一为不区分大小写
                        pattern = re.compile(re.escape(placeholder), re.IGNORECASE)
                        text = pattern.sub(str(first_data_row[i]), text)
        
        return text
    
    def batch_create_drafts(self):
        """批量创建草稿（为Excel每行创建一封邮件）"""
        # 检查是否需要动态询问Excel文件
        if self.prompt_excel_var.get():
            if messagebox.askyesno("动态Excel模式", "您开启了'运行时询问Excel文件'。\n是否选择新的Excel数据源？"):
                # 1. 选择文件
                self.select_excel_file()
                if not self.excel_path_var.get():
                    return # 用户取消
                
                # 2. 尝试自动检测并读取
                # 如果sheet名为空，默认选第一个
                if not self.sheet_combo.get() and self.sheet_combo['values']:
                    self.sheet_combo.current(0)
                
                # 尝试读取数据
                self.read_excel_data()
        
        if not self.current_excel_data or len(self.current_excel_data) < 2:
            messagebox.showwarning("警告", "请先读取Excel数据，且数据至少需要2行（表头+数据）")
            return
        
        # 从列表框获取邮箱地址
        to_list = [x.strip() for x in self.to_entry_var.get().split(';') if x.strip()]
        cc_list = [x.strip() for x in self.cc_entry_var.get().split(';') if x.strip()]
        bcc_list = [x.strip() for x in self.bcc_entry_var.get().split(';') if x.strip()]
        
        subject_template = self.subject_var.get().strip()
        body_template = self.body_text.get(1.0, tk.END).strip()
        
        # 检查是否有收件人或使用了变量
        to_str = "; ".join(to_list) if to_list else ""
        has_variable = "{" in to_str or not to_list
        
        if not has_variable and not to_list:
            messagebox.showwarning("警告", "批量模式需要添加收件人或在收件人中使用变量，如 {邮箱}")
            return
        
        if not subject_template:
            messagebox.showwarning("警告", "请填写主题")
            return
        
        # 确认操作
        data_count = len(self.current_excel_data) - 1  # 减去表头
        if not messagebox.askyesno("确认批量创建", 
                                    f"将为 {data_count} 行数据创建 {data_count} 封草稿邮件。\n确定继续？"):
            return
        
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            headers = self.current_excel_data[0]
            created_count = 0
            
            # 智能匹配逻辑 - 用于 {SMART_IMAGE} 和数据源
            processed_excel_paths = self.process_template_variables(self.excel_path_var.get())
            current_excel_path = processed_excel_paths.split(';')[0]
            sheet_name = self.sheet_combo.get()
            data_range = self.range_var.get()

            # --- 全流程数据一致性处理 ---
            # 如果开启了智能匹配，或即使没开启，为了确保数据和截图的一致性，
            # 这里强制使用解析后的路径重新读取数据。
            
            if self.smart_match_var.get() and current_excel_path:
                latest_path = self.get_latest_file_in_folder(current_excel_path)
                if latest_path != current_excel_path:
                    print(f"Batch Create: Switching {current_excel_path} -> {latest_path}")
                    current_excel_path = latest_path
            
            # 自动重新读取数据，确保 Self.current_excel_data 是最新的
            # 并且来自正确的 current_excel_path
            if current_excel_path and os.path.exists(current_excel_path):
                 print(f"Reloading data from: {current_excel_path} for consistency...")
                 result, error = self._internal_read_excel(current_excel_path, sheet_name, data_range)
                 if result:
                     rows, _ = result
                     if rows and len(rows) > 1:
                         self.current_excel_data = rows
                         headers = rows[0] # 更新表头
                     else:
                         print("Warning: Reloaded data is empty or insufficient.")
                 else:
                     print(f"Warning: Failed to reload data: {error}")


            # 预先处理 {SMART_IMAGE}
            # 这会生成一个通用的截图用于所有批量邮件
            smart_img_info = None
            if "{SMART_IMAGE}" in body_template:
                body_template, smart_img_info = self.process_smart_image_logic(
                    body_template, current_excel_path, sheet_name, data_range
                )

            # 预先生成图片（如果需要且是静态的）
            # 这里我们假设批量邮件使用相同的Excel截图（来自配置的范围）
            batch_images = []
            if "{EXCEL_IMAGE}" in body_template:
                 try:
                     batch_images = self.generate_excel_images()
                 except Exception as e:
                     print(f"批量生成图片失败: {e}")

            # 从第二行开始（跳过表头）
            data_rows = self.current_excel_data[1:]
            total = len(data_rows)
            
            # 创建进度窗口
            progress_win = tk.Toplevel(self.root)
            progress_win.title("批量创建进度")
            progress_win.geometry("400x120")
            progress_win.transient(self.root)
            progress_win.grab_set()
            progress_win.resizable(False, False)
            
            progress_label = ttk.Label(progress_win, text="正在创建草稿...", padding=10)
            progress_label.pack()
            progress_bar = ttk.Progressbar(progress_win, length=350, mode='determinate', maximum=total)
            progress_bar.pack(padx=20, pady=5)
            progress_detail = ttk.Label(progress_win, text=f"0 / {total}")
            progress_detail.pack()
            
            for idx, row_data in enumerate(data_rows):
                # 处理收件人
                to_filled_list = [self.process_template_variables(email, row_data, headers) for email in to_list]
                cc_filled_list = [self.process_template_variables(email, row_data, headers) for email in cc_list]
                bcc_filled_list = [self.process_template_variables(email, row_data, headers) for email in bcc_list]
                
                # 替换主题和正文中的变量
                subject_filled = self.process_template_variables(subject_template, row_data, headers)
                body_filled = self.process_template_variables(body_template, row_data, headers)
                
                # 创建邮件
                mail = outlook.CreateItem(0)
                mail.To = "; ".join(to_filled_list) if to_filled_list else to_str
                if cc_filled_list:
                    mail.CC = "; ".join(cc_filled_list)
                if bcc_filled_list:
                    mail.BCC = "; ".join(bcc_filled_list)
                mail.Subject = subject_filled
                
                # 处理HTML正文
                body_html = body_filled
                
                # 处理 {EXCEL_IMAGE} 占位符
                current_email_inline_images = []
                if "{EXCEL_IMAGE}" in body_html and batch_images:
                    img_fragments = []
                    for img_path in batch_images:
                        cid = f"img_{uuid.uuid4().hex}@local"
                        current_email_inline_images.append({'path': img_path, 'cid': cid})
                        img_tag = f'<img src="cid:{cid}" alt="Excel表格" style="max-width:100%; border:1px solid #ccc;"><br/>'
                        img_fragments.append(img_tag)
                    
                    body_html = body_html.replace("{EXCEL_IMAGE}", "\n".join(img_fragments))

                body_html = body_html.replace("\n", "<br>")
                mail.HTMLBody = body_html
                
                # 添加附件
                for attachment_path in self.attachments:
                    if os.path.exists(attachment_path):
                        mail.Attachments.Add(attachment_path)
                
                # 添加内嵌图片
                for img in current_email_inline_images:
                    path = img['path']
                    cid = img['cid']
                    if os.path.exists(path):
                        try:
                            att = mail.Attachments.Add(path)
                            att.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)
                        except:
                            pass
                
                # 添加 {SMART_IMAGE} 图片
                if smart_img_info:
                    path = smart_img_info['path']
                    cid = smart_img_info['cid']
                    if os.path.exists(path):
                        try:
                            att = mail.Attachments.Add(path)
                            att.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)
                        except:
                            pass
                
                mail.Save()
                created_count += 1
                
                # 更新进度
                progress_bar['value'] = idx + 1
                progress_detail.config(text=f"{idx + 1} / {total}")
                progress_win.update()
            
            # 关闭进度窗口
            progress_win.destroy()
            
            # 清理智能截图临时文件
            if smart_img_info and os.path.exists(smart_img_info['path']):
                try:
                    os.remove(smart_img_info['path'])
                except:
                    pass

            messagebox.showinfo("成功", f"成功创建 {created_count} 封草稿邮件！")
            self.status_var.set(f"批量创建完成 - {created_count} 封草稿")
            
        except Exception as e:
            messagebox.showerror("错误", f"批量创建失败: {str(e)}")
    
    def preview_excel_data(self):
        """预览Excel数据"""
        if not self.current_excel_data:
            messagebox.showwarning("警告", "请先读取Excel数据")
            return
        
        # 创建预览窗口
        preview_window = tk.Toplevel(self.root)
        preview_window.title("Excel 数据预览")
        preview_window.geometry("800x600")
        
        # 创建Treeview显示表格
        frame = ttk.Frame(preview_window, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 滚动条
        scrollbar_y = ttk.Scrollbar(frame, orient=tk.VERTICAL)
        scrollbar_x = ttk.Scrollbar(frame, orient=tk.HORIZONTAL)
        
        # 创建Treeview
        if self.current_excel_data and len(self.current_excel_data) > 0:
            headers = [str(h) for h in self.current_excel_data[0]]
            tree = ttk.Treeview(frame, columns=headers, show='headings',
                               yscrollcommand=scrollbar_y.set,
                               xscrollcommand=scrollbar_x.set)
            
            # 设置列标题
            for header in headers:
                tree.heading(header, text=header)
                tree.column(header, width=100)
            
            # 添加数据行
            for row in self.current_excel_data[1:]:
                tree.insert('', tk.END, values=row)
            
            scrollbar_y.config(command=tree.yview)
            scrollbar_x.config(command=tree.xview)
            
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
            scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
            
            ttk.Label(preview_window, text=f"共 {len(self.current_excel_data)-1} 行数据").pack(pady=5)
    
    def export_config(self):
        """导出配置到文件"""
        config_name = self.config_combo.get()
        if not config_name:
            messagebox.showwarning("警告", "请先选择要导出的配置")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="导出配置",
            defaultextension=".json",
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
            initialfile=f"{config_name}.json"
        )
        
        if file_path:
            try:
                export_data = {config_name: self.configs[config_name]}
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
                messagebox.showinfo("成功", f"配置已导出到：\n{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def import_config(self):
        """从文件导入配置"""
        file_path = filedialog.askopenfilename(
            title="导入配置",
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    imported_configs = json.load(f)
                
                imported_count = 0
                for name, config in imported_configs.items():
                    if name in self.configs:
                        if not messagebox.askyesno("确认覆盖", 
                                                   f"配置 '{name}' 已存在，是否覆盖？"):
                            continue
                    self.configs[name] = config
                    imported_count += 1
                
                if self.save_configs():
                    self.update_config_list()
                    messagebox.showinfo("成功", f"成功导入 {imported_count} 个配置")
                    
            except Exception as e:
                messagebox.showerror("错误", f"导入失败: {str(e)}")
    
    def save_to_history(self, to, subject):
        """保存到历史记录"""
        try:
            history = []
            if os.path.exists(self.history_file):
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
            
            history.append({
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "to": to,
                "subject": subject
            })
            
            # 只保留最近100条记录
            history = history[-100:]
            
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"保存历史记录失败: {e}")
    
    def start_auto_save(self):
        """启动自动保存"""
        # 每60秒自动保存一次当前配置
        if self.config_name_var.get().strip():
            # 静默保存（不显示提示）
            config_name = self.config_name_var.get().strip()
            if config_name in self.configs:
                to_list = [x.strip() for x in self.to_entry_var.get().split(';') if x.strip()]
                cc_list = [x.strip() for x in self.cc_entry_var.get().split(';') if x.strip()]
                bcc_list = [x.strip() for x in self.bcc_entry_var.get().split(';') if x.strip()]
                config_data = {
                    "excel_path": self.excel_path_var.get(),
                    "sheet_name": self.sheet_combo.get(),
                    "data_range": self.range_var.get(),
                    "to": to_list,
                    "cc": cc_list,
                    "bcc": bcc_list,
                    "subject": self.subject_var.get(),
                    "body": self.body_text.get(1.0, tk.END).strip(),
                    "attachments": self.attachments.copy(),
                    "inline_images": self.inline_images.copy() if hasattr(self, 'inline_images') else [],
                    "custom_placeholders": self.custom_placeholders.copy(),
                    "high_priority": self.priority_var.get(),
                    "read_receipt": self.receipt_var.get(),
                    "smart_match": self.smart_match_var.get(),
                    "filename_keyword": self.filename_keyword_var.get(),
                    "attach_excel": self.attach_excel_var.get()
                }
                self.configs[config_name] = config_data
                self.save_configs()
        
        # 继续定时器
        self.auto_save_timer = self.root.after(60000, self.start_auto_save)
    
    def select_multiple_configs(self):
        """选择多个配置进行批量处理"""
        if not self.configs:
            messagebox.showwarning("警告", "没有可用的配置")
            return
        
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title("选择要批量处理的配置")
        select_window.geometry("500x600")
        select_window.transient(self.root)
        select_window.grab_set()
        
        # 说明标签
        info_frame = ttk.Frame(select_window, padding="10")
        info_frame.pack(fill=tk.X)
        ttk.Label(info_frame, text="选择要批量创建草稿的配置（可多选）：", 
                 font=('TkDefaultFont', 10, 'bold')).pack(anchor=tk.W)
        ttk.Label(info_frame, text="勾选的配置将依次创建草稿邮件", 
                 font=('TkDefaultFont', 9)).pack(anchor=tk.W, pady=(0, 5))
        
        # 创建可滚动的复选框列表
        list_frame = ttk.Frame(select_window, padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(list_frame)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 存储复选框变量
        checkbox_vars = {}
        
        # 为每个配置创建复选框
        for i, config_name in enumerate(sorted(self.configs.keys())):
            var = tk.BooleanVar(value=config_name in self.selected_configs)
            checkbox_vars[config_name] = var
            
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, padx=5, pady=2)
            
            cb = ttk.Checkbutton(frame, text=config_name, variable=var)
            cb.pack(side=tk.LEFT)
            
            # 显示配置详情
            config = self.configs[config_name]
            to_raw = config.get('to', '')
            to_str = '; '.join(to_raw) if isinstance(to_raw, list) else str(to_raw)
            to = to_str[:30] + '...' if len(to_str) > 30 else to_str
            subject_str = config.get('subject', '')
            subject = subject_str[:40] + '...' if len(subject_str) > 40 else subject_str
            
            detail = f"收件人: {to} | 主题: {subject}"
            ttk.Label(frame, text=detail, font=('TkDefaultFont', 8), foreground='gray').pack(side=tk.LEFT, padx=(10, 0))
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 统计信息
        stats_frame = ttk.Frame(select_window, padding="10")
        stats_frame.pack(fill=tk.X)
        stats_label = ttk.Label(stats_frame, text=f"总配置数: {len(self.configs)}", 
                                font=('TkDefaultFont', 9))
        stats_label.pack()
        
        # 按钮区域
        button_frame = ttk.Frame(select_window, padding="10")
        button_frame.pack(fill=tk.X)
        
        def select_all():
            for var in checkbox_vars.values():
                var.set(True)
        
        def select_none():
            for var in checkbox_vars.values():
                var.set(False)
        
        def confirm_selection():
            self.selected_configs = [name for name, var in checkbox_vars.items() if var.get()]
            if not self.selected_configs:
                messagebox.showwarning("警告", "请至少选择一个配置")
                return
            
            messagebox.showinfo("已选择", f"已选择 {len(self.selected_configs)} 个配置：\n" + 
                              "\n".join(f"• {name}" for name in self.selected_configs[:10]) +
                              (f"\n... 还有 {len(self.selected_configs)-10} 个" if len(self.selected_configs) > 10 else ""))
            select_window.destroy()
        
        ttk.Button(button_frame, text="全选", command=select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="全不选", command=select_none).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="确定", command=confirm_selection).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="取消", command=select_window.destroy).pack(side=tk.RIGHT, padx=5)
    
    def batch_create_all_drafts(self):
        """一键批量创建所有选中配置的草稿"""
        if not self.selected_configs:
            # 如果没有选择，弹出选择窗口
            result = messagebox.askyesno("提示", "尚未选择配置。\n是否现在选择要批量处理的配置？")
            if result:
                self.select_multiple_configs()
            return
        
        # 确认操作
        confirm_msg = f"将为以下 {len(self.selected_configs)} 个配置创建草稿邮件：\n\n"
        confirm_msg += "\n".join(f"  {i+1}. {name}" for i, name in enumerate(self.selected_configs[:10]))
        if len(self.selected_configs) > 10:
            confirm_msg += f"\n  ... 还有 {len(self.selected_configs)-10} 个配置"
        confirm_msg += "\n\n确定继续？"
        
        if not messagebox.askyesno("确认批量创建", confirm_msg):
            return
        
        # 创建进度窗口
        progress_window = tk.Toplevel(self.root)
        progress_window.title("批量创建进度")
        progress_window.geometry("600x400")
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # 进度信息
        info_frame = ttk.Frame(progress_window, padding="10")
        info_frame.pack(fill=tk.X)
        
        progress_label = ttk.Label(info_frame, text="正在处理...", font=('TkDefaultFont', 10, 'bold'))
        progress_label.pack(anchor=tk.W)
        
        progress_bar = ttk.Progressbar(info_frame, length=560, mode='determinate')
        progress_bar.pack(fill=tk.X, pady=10)
        
        status_label = ttk.Label(info_frame, text="", font=('TkDefaultFont', 9))
        status_label.pack(anchor=tk.W)
        
        # 日志区域
        log_frame = ttk.LabelFrame(progress_window, text="处理日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        log_text = scrolledtext.ScrolledText(log_frame, width=70, height=15, wrap=tk.WORD)
        log_text.pack(fill=tk.BOTH, expand=True)
        
        def log_message(msg, level="INFO"):
            timestamp = datetime.now().strftime("%H:%M:%S")
            color_tag = "error" if level == "ERROR" else "success" if level == "SUCCESS" else "info"
            log_text.insert(tk.END, f"[{timestamp}] ", "timestamp")
            log_text.insert(tk.END, f"{msg}\n", color_tag)
            log_text.see(tk.END)
            log_text.update()
        
        # 配置颜色标签
        log_text.tag_config("timestamp", foreground="gray")
        log_text.tag_config("info", foreground="black")
        log_text.tag_config("success", foreground="green")
        log_text.tag_config("error", foreground="red")
        
        # 开始处理
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            log_message("✓ Outlook 已连接", "SUCCESS")
            
            total = len(self.selected_configs)
            success_count = 0
            error_count = 0
            
            progress_bar['maximum'] = total
            
            for index, config_name in enumerate(self.selected_configs, 1):
                try:
                    progress_label.config(text=f"正在处理: {config_name} ({index}/{total})")
                    status_label.config(text=f"进度: {index}/{total}")
                    progress_bar['value'] = index
                    
                    log_message(f"开始处理配置: {config_name}")
                    
                    config = self.configs[config_name]
                    
                    # 加载配置数据（支持列表和字符串两种格式）
                    to_data = config.get('to', [])
                    cc_data = config.get('cc', [])
                    bcc_data = config.get('bcc', [])
                    
                    # 兼容旧格式（字符串）转换为列表
                    if isinstance(to_data, str):
                        to_list = [e.strip() for e in to_data.split(';') if e.strip()]
                    else:
                        to_list = to_data
                    
                    if isinstance(cc_data, str):
                        cc_list = [e.strip() for e in cc_data.split(';') if e.strip()]
                    else:
                        cc_list = cc_data
                    
                    if isinstance(bcc_data, str):
                        bcc_list = [e.strip() for e in bcc_data.split(';') if e.strip()]
                    else:
                        bcc_list = bcc_data
                    
                    subject = config.get('subject', '').strip()
                    body = config.get('body', '').strip()
                    attachments = config.get('attachments', [])
                    inline_images_config = config.get('inline_images', []) # 获取配置中的内嵌图片
                    
                    # 检查必填字段
                    if not to_list:
                        log_message(f"  ⚠ 跳过（缺少收件人）", "ERROR")
                        error_count += 1
                        continue
                    
                    if not subject:
                        log_message(f"  ⚠ 跳过（缺少主题）", "ERROR")
                        error_count += 1
                        continue
                    
                    # 读取Excel数据（如果有）
                    # 方案一：智能匹配最新文件
                    smart_match = config.get("smart_match", False)
                    filename_keyword = config.get("filename_keyword", "") # 获取该配置的关键词
                    excel_path_raw = config.get('excel_path', '')
                    sheet_name = config.get('sheet_name', '')
                    data_range = config.get('data_range', '')
                    
                    # 确保提取单个路径 (以防配置中保存了多个以分号分隔的路径)
                    excel_path = excel_path_raw.split(';')[0].strip() if excel_path_raw else ""

                    # === 修复：预处理 Excel 路径中的占位符 ===
                    # 必须在智能匹配和读取之前解析路径中的 {year}, {data} 等变量
                    if excel_path:
                        # 1. 准备变量字典 (合并全局和配置特有的)
                        path_vars = {
                            "{date}": datetime.now().strftime("%Y-%m-%d"),
                            "{time}": datetime.now().strftime("%H:%M:%S"),
                            "{datetime}": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        
                        # 全局自定义变量
                        for k, v in self.custom_placeholders.items():
                             path_vars["{" + str(k) + "}"] = str(v)
                             
                        # 配置特定变量
                        config_spec_vars = config.get("custom_placeholders", {})
                        for k, v in config_spec_vars.items():
                            path_vars["{" + str(k) + "}"] = str(v)
                        
                        # 2. 执行替换 (不区分大小写)
                        for k, v in path_vars.items():
                            excel_path = re.sub(re.escape(k), str(v), excel_path, flags=re.IGNORECASE)

                    if smart_match and excel_path:
                        # 传入具体的关键词，而不是使用 UI 全局变量
                        latest_path = self.get_latest_file_in_folder(excel_path, keyword=filename_keyword)
                        if latest_path and latest_path != excel_path:
                            log_message(f"  ⚡ 智能匹配: {os.path.basename(latest_path)}")
                            excel_path = latest_path

                    # 方案二：处理 {SMART_IMAGE} (在读取 Excel 数据前处理)
                    temp_inline_images = []
                    if "{SMART_IMAGE}" in body:
                        log_message(f"  🖼 正在生成智能截图...")
                        body, smart_img_info = self.process_smart_image_logic(
                            body, excel_path, sheet_name, data_range
                        )
                        if smart_img_info:
                            temp_inline_images.append(smart_img_info)
                        else:
                            log_message(f"  ⚠ 智能截图生成失败", "ERROR")
                    
                    excel_data = None
                    if excel_path and os.path.exists(excel_path) and sheet_name and data_range:
                        try:
                            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                            ws = wb[sheet_name]
                            cells = ws[data_range]
                            
                            data_rows = []
                            if isinstance(cells, tuple) and len(cells) > 0:
                                if isinstance(cells[0], tuple):
                                    for row in cells:
                                        row_data = [cell.value if cell.value is not None else "" for cell in row]
                                        data_rows.append(row_data)
                                else:
                                    row_data = [cell.value if cell.value is not None else "" for cell in cells]
                                    data_rows.append(row_data)
                            else:
                                data_rows.append([cells.value if cells.value is not None else ""])
                            
                            wb.close()
                            excel_data = data_rows
                            log_message(f"  ✓ Excel数据已加载 ({len(data_rows)} 行)")
                        except Exception as e:
                            log_message(f"  ⚠ Excel读取失败: {str(e)}", "ERROR")
                    
                    # === 核心修改：确保每个配置只生成一个草稿 ===
                    # 无论Excel有多少行数据，只取第一行数据（如果有）来替换变量
                    # 并生成一封邮件，而不是进行邮件合并操作
                    
                    # 1. 准备变量字典
                    variables = {}
                    # 添加Excel第一行数据的变量
                    if excel_data and len(excel_data) > 1:
                        headers = excel_data[0]
                        first_row = excel_data[1]
                        for i in range(min(len(headers), len(first_row))):
                            variables["{" + str(headers[i]) + "}"] = str(first_row[i])
                    
                    # 添加自定义占位符
                    for k, v in self.custom_placeholders.items():
                        variables["{" + str(k) + "}"] = str(v)
                        
                    config_vars = config.get("custom_placeholders", {})
                    for k, v in config_vars.items():
                        variables["{" + str(k) + "}"] = str(v)

                    # 2. 替换变量 (主题和正文)
                    subject_filled = subject
                    body_filled = body

                    # 处理收件人 (支持变量)
                    to_filled_list = []
                    for email in to_list:
                        email_filled = email
                        for k, v in variables.items():
                             email_filled = email_filled.replace(k, str(v))
                        to_filled_list.append(email_filled)
                    
                    # 替换一般变量
                    for k, v in variables.items():
                        subject_filled = subject_filled.replace(k, str(v))
                        body_filled = body_filled.replace(k, str(v))
                    
                    # 处理时间日期变量
                    curr_time = datetime.now()
                    subject_filled = subject_filled.replace("{DATE}", curr_time.strftime("%Y-%m-%d"))
                    subject_filled = subject_filled.replace("{TIME}", curr_time.strftime("%H:%M:%S"))
                    body_filled = body_filled.replace("{DATE}", curr_time.strftime("%Y-%m-%d"))
                    body_filled = body_filled.replace("{TIME}", curr_time.strftime("%H:%M:%S"))

                    # 构建邮件
                    mail = outlook.CreateItem(0)
                    mail.To = "; ".join(to_filled_list)
                    if cc_list:
                         mail.CC = "; ".join(cc_list)
                    if bcc_list:
                         mail.BCC = "; ".join(bcc_list)
                    
                    if config.get("high_priority", False):
                        mail.Importance = 2
                    if config.get("read_receipt", False):
                        mail.ReadReceiptRequested = True

                    mail.Subject = subject_filled
                    
                    # 3. 处理正文特殊内容
                    body_html = body_filled
                    
                    # 处理 {EXCEL_DATA} 表格插入
                    if "{EXCEL_DATA}" in body_html and excel_data:
                        # 尝试使用内置格式化方法
                        try:
                            html_table = self.generate_formatted_html_table(excel_data)
                            body_html = body_html.replace("{EXCEL_DATA}", html_table)
                        except:
                            # 失败则使用简单表格
                            html_table = "<table border='1'><tr>" + "</tr><tr>".join(["".join([f"<td>{c}</td>" for c in r]) for r in excel_data]) + "</tr></table>"
                            body_html = body_html.replace("{EXCEL_DATA}", html_table)
                    
                    # 处理 {EXCEL_IMAGE}
                    # 批量模式下暂不支持动态生成 Excel 截图，除非配置中已有
                    # 这里主要依靠之前的 self.process_template_variables 逻辑或静态添加
                    
                    # 替换换行
                    body_html = body_html.replace("\n", "<br>")
                    
                    mail.HTMLBody = body_html
                    
                    # 4. 添加附件和内嵌图片
                    
                    # 准备附件列表
                    final_attachments = attachments.copy()
                    # 如果配置要求附带源文件，且文件存在
                    if config.get("attach_excel", False) and excel_path and os.path.exists(excel_path):
                        if excel_path not in final_attachments:
                            final_attachments.append(excel_path)
                            
                    # 普通附件
                    for att_path in final_attachments:
                        if os.path.exists(att_path):
                            mail.Attachments.Add(att_path)
                            
                    # 内嵌图片 (Config自带 + Smart Image)
                    combined_inline_pics = inline_images_config + temp_inline_images
                    for img in combined_inline_pics:
                        path = img.get('path')
                        cid = img.get('cid')
                        if path and cid and os.path.exists(path):
                            try:
                                att = mail.Attachments.Add(path)
                                att.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)
                            except:
                                pass
                                
                    mail.Save()
                    
                    log_message(f"  ✓ 已创建草稿: {subject_filled}")
                    success_count += 1

                    # 清理本次配置生成的临时截图
                    for img in temp_inline_images:
                        try:
                            if os.path.exists(img['path']):
                                os.remove(img['path'])
                        except:
                            pass

                except Exception as e:
                    log_message(f"  ❌ 处理失败: {str(e)}", "ERROR")
                    error_count += 1
                    import traceback
                    print(traceback.format_exc())
            
            progress_bar['value'] = total
            status_label.config(text="完成")
            
            # 完成
            progress_label.config(text="✓ 批量处理完成！")
            log_message("=" * 50)
            log_message(f"处理完成！成功: {success_count}, 失败: {error_count}", "SUCCESS" if error_count == 0 else "INFO")
            
            # 关闭按钮
            close_button = ttk.Button(progress_window, text="关闭", command=progress_window.destroy)
            close_button.pack(pady=10)
            
            # 更新状态栏
            self.status_var.set(f"批量创建完成 - 成功: {success_count}, 失败: {error_count}")
            
        except Exception as e:
            log_message(f"严重错误: {str(e)}", "ERROR")
            messagebox.showerror("错误", f"批量创建失败: {str(e)}")
            progress_window.destroy()

    def preview_batch_merge(self):
        """预览批量合并效果（显示第一行数据的合并结果）"""
        if not self.current_excel_data or len(self.current_excel_data) < 2:
            messagebox.showwarning("警告", "请先读取Excel数据，且数据至少需要2行（表头+数据）")
            return
        
        # 获取第一行数据
        headers = self.current_excel_data[0]
        first_row = self.current_excel_data[1]
        
        # 创建变量字典
        variables = {"{" + str(headers[i]) + "}": str(first_row[i]) 
                    for i in range(min(len(headers), len(first_row)))}
        
        # 获取模板
        subject_template = self.subject_var.get().strip()
        body_template = self.body_text.get(1.0, tk.END).strip()
        
        # 替换变量
        subject_filled = subject_template
        body_filled = body_template
        
        for placeholder, value in variables.items():
            subject_filled = subject_filled.replace(placeholder, value)
            body_filled = body_filled.replace(placeholder, value)
            
        # 处理其他变量
        subject_filled = self.process_template_variables(subject_filled)
        body_filled = self.process_template_variables(body_filled)
        
        # 显示预览窗口
        preview_win = tk.Toplevel(self.root)
        preview_win.title("批量合并预览 (第1行数据)")
        preview_win.geometry("800x600")
        
        # 提示信息
        ttk.Label(preview_win, text="这是使用Excel第一行数据生成的预览效果：", 
                 font=('TkDefaultFont', 10, 'bold'), foreground='blue').pack(pady=10)
        
        # 显示变量映射
        var_frame = ttk.LabelFrame(preview_win, text="变量映射", padding="5")
        var_frame.pack(fill=tk.X, padx=10, pady=5)
        
        var_text = " | ".join([f"{k}={v}" for k, v in list(variables.items())[:5]])
        if len(variables) > 5:
            var_text += " ..."
        ttk.Label(var_frame, text=var_text, font=('TkDefaultFont', 9)).pack(anchor=tk.W)
        
        # 预览内容
        content_frame = ttk.LabelFrame(preview_win, text="邮件内容", padding="10")
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        ttk.Label(content_frame, text=f"主题: {subject_filled}", font=('TkDefaultFont', 10, 'bold')).pack(anchor=tk.W, pady=(0, 10))
        ttk.Separator(content_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(0, 10))
        
        text_widget = scrolledtext.ScrolledText(content_frame, wrap=tk.WORD, font=('TkDefaultFont', 10))
        text_widget.pack(fill=tk.BOTH, expand=True)
        text_widget.insert(1.0, body_filled)
        text_widget.config(state='disabled')
        
        ttk.Button(preview_win, text="关闭", command=preview_win.destroy).pack(pady=10)
    
    def preview_all_configs(self):
        """预览所有选中的配置"""
        if not self.selected_configs:
            messagebox.showwarning("警告", "请先选择要预览的配置")
            return
        
        # 创建预览窗口
        preview_window = tk.Toplevel(self.root)
        preview_window.title(f"批量配置预览 ({len(self.selected_configs)} 个)")
        preview_window.geometry("900x700")
        preview_window.transient(self.root)
        
        # 创建笔记本（标签页）
        notebook = ttk.Notebook(preview_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        for config_name in self.selected_configs:
            config = self.configs[config_name]
            
            # 为每个配置创建一个标签页
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=config_name[:20])
            
            # 显示配置详情
            text_widget = scrolledtext.ScrolledText(frame, wrap=tk.WORD, width=80, height=35)
            text_widget.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # 处理收件人数据（兼容列表和字符串格式）
            to_data = config.get('to', [])
            cc_data = config.get('cc', [])
            bcc_data = config.get('bcc', [])
            
            if isinstance(to_data, list):
                to_str = "; ".join(to_data)
            else:
                to_str = to_data
            
            if isinstance(cc_data, list):
                cc_str = "; ".join(cc_data)
            else:
                cc_str = cc_data
            
            if isinstance(bcc_data, list):
                bcc_str = "; ".join(bcc_data)
            else:
                bcc_str = bcc_data
            
            preview_text = f"配置名称: {config_name}\n"
            preview_text += "=" * 60 + "\n\n"
            preview_text += f"收件人 (To): {to_str}\n"
            preview_text += f"抄送 (CC): {cc_str}\n"
            preview_text += f"密送 (BCC): {bcc_str}\n\n"
            preview_text += f"主题: {config.get('subject', '')}\n"
            preview_text += "-" * 60 + "\n\n"
            preview_text += "正文:\n"
            preview_text += config.get('body', '')
            preview_text += "\n\n" + "-" * 60 + "\n\n"
            
            if config.get('excel_path'):
                preview_text += f"Excel文件: {config.get('excel_path', '')}\n"
                preview_text += f"工作表: {config.get('sheet_name', '')}\n"
                preview_text += f"数据范围: {config.get('data_range', '')}\n\n"
            
            if config.get('attachments'):
                preview_text += "附件:\n"
                for att in config.get('attachments', []):
                    preview_text += f"  • {os.path.basename(att)}\n"
            
            text_widget.insert(1.0, preview_text)
            text_widget.config(state='disabled')
        
        # 关闭按钮
        ttk.Button(preview_window, text="关闭", command=preview_window.destroy).pack(pady=5)
    
    def setup_config_browser(self, parent):
        """设置右侧配置浏览器"""
        browser_frame = ttk.LabelFrame(parent, text="📚 已保存配置", padding="10")
        # 添加到 PanedWindow
        parent.add(browser_frame, weight=1)
        
        browser_frame.rowconfigure(1, weight=1)
        browser_frame.columnconfigure(0, weight=1)
        
        # 搜索框
        search_frame = ttk.Frame(browser_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(search_frame, text="🔍").pack(side=tk.LEFT)
        self.config_search_var = tk.StringVar()
        self.config_search_var.trace('w', self.filter_config_list)
        search_entry = ttk.Entry(search_frame, textvariable=self.config_search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        # 配置列表 (使用 Treeview 替代 Listbox 以显示更多信息)
        list_frame = ttk.Frame(browser_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        h_scrollbar = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL)
        
        self.config_tree = ttk.Treeview(list_frame, columns=("name", "file"), show="headings", 
                                      yscrollcommand=scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.config_tree.heading("name", text="配置名称")
        self.config_tree.heading("file", text="Excel 文件")
        
        self.config_tree.column("name", width=150, minwidth=100)
        self.config_tree.column("file", width=200, minwidth=100)
        
        scrollbar.config(command=self.config_tree.yview)
        h_scrollbar.config(command=self.config_tree.xview)
        
        self.config_tree.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.E, tk.W))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.E, tk.W))
        
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        
        # 绑定双击事件加载配置
        self.config_tree.bind('<Double-Button-1>', self.load_config_from_browser)
        self.config_tree.bind('<Return>', self.load_config_from_browser)
        
        # 按钮区域
        button_frame = ttk.Frame(browser_frame)
        button_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Button(button_frame, text="加载", command=lambda: self.load_config_from_browser(None), 
                  width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="删除", command=self.delete_from_browser, 
                  width=10).pack(side=tk.LEFT, padx=2)
        
        # 信息标签
        self.browser_info_label = ttk.Label(browser_frame, text="", font=('TkDefaultFont', 8))
        self.browser_info_label.pack(pady=(5, 0))
        
        # 初始化列表
        self.update_config_browser()
    
    def update_config_browser(self):
        """更新配置浏览器列表"""
        # 清空现有项
        for item in self.config_tree.get_children():
            self.config_tree.delete(item)
        
        search_term = self.config_search_var.get().lower() if hasattr(self, 'config_search_var') else ""
        
        config_list = sorted(self.configs.keys())
        count = 0
        
        for name in config_list:
            if search_term and search_term not in name.lower():
                continue
                
            config = self.configs[name]
            excel_path = config.get("excel_path", "")
            # 如果有多个文件，显示数量
            if ";" in excel_path:
                file_display = f"[{excel_path.count(';')+1} 个文件] {os.path.basename(excel_path.split(';')[0])}..."
            else:
                file_display = os.path.basename(excel_path) if excel_path else "(无)"
                # 如果文件名相同但路径不同，显示部分路径以区分
                # 这里简单处理：显示文件名，鼠标悬停或选中时可以看到详情（虽然Treeview没有默认tooltip）
                # 实际上，用户可以通过查看 file_display 来区分
            
            # 插入数据，values对应 columns
            self.config_tree.insert("", tk.END, values=(name, file_display), tags=(name,))
            count += 1
            
        if hasattr(self, 'browser_info_label'):
            self.browser_info_label.config(text=f"共 {count} 个配置")

    def filter_config_list(self, *args):
        """过滤配置列表"""
        self.update_config_browser()

    def load_config_from_browser(self, event):
        """从浏览器加载配置"""
        selection = self.config_tree.selection()
        if not selection:
            return
            
        item = selection[0]
        values = self.config_tree.item(item, "values")
        if values:
            config_name = values[0]
            self.config_combo.set(config_name)
            self.load_selected_config()

    def delete_from_browser(self):
        """从浏览器删除配置"""
        selection = self.config_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要删除的配置")
            return
            
        item = selection[0]
        values = self.config_tree.item(item, "values")
        if values:
            config_name = values[0]
            if messagebox.askyesno("确认删除", f"确定要删除配置 '{config_name}' 吗？"):
                del self.configs[config_name]
                if self.save_configs():
                    self.update_config_list() # 这会触发 update_config_browser
                    messagebox.showinfo("成功", f"配置 '{config_name}' 已删除")

    
    def add_manual_email_dialog(self, entry_widget):
        """弹出高级对话框编辑/添加邮箱（支持多行粘贴）"""
        # 创建弹窗
        dialog = tk.Toplevel(self.root)
        dialog.title("批量编辑/添加收件人")
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry('{}x{}+{}+{}'.format(width, height, x, y))
        
        # 说明
        instruction_frame = ttk.Frame(dialog, padding="10 10 10 5")
        instruction_frame.pack(fill=tk.X)
        ttk.Label(instruction_frame, text="在此处输入或粘贴邮箱地址:", font=('TkDefaultFont', 10, 'bold')).pack(anchor=tk.W)
        ttk.Label(instruction_frame, text="• 每行一个地址，或使用分号/逗号分隔\n• 支持直接从Excel列复制粘贴", 
                 foreground="gray").pack(anchor=tk.W, pady=(2, 0))
        
        # 文本框
        text_frame = ttk.Frame(dialog, padding="10 0 10 5")
        text_frame.pack(fill=tk.BOTH, expand=True)
        text_area = scrolledtext.ScrolledText(text_frame, width=50, height=10)
        text_area.pack(fill=tk.BOTH, expand=True)
        
        # 加载现有内容
        current_text = entry_widget.get().strip()
        if current_text:
            # 智能分割：尝试分号、逗号、换行
            raw_items = re.split(r'[;,\n]', current_text)
            # 清理空白项并去重(保持顺序)
            clean_items = []
            seen = set()
            for item in raw_items:
                i = item.strip()
                if i and i not in seen:
                    clean_items.append(i)
                    seen.add(i)
            
            text_area.insert(1.0, "\n".join(clean_items))
        
        # 按钮区域
        btn_frame = ttk.Frame(dialog, padding="10")
        btn_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        def save():
            content = text_area.get(1.0, tk.END).strip()
            if content:
                # 统一分隔符处理
                # 1. 替换中文分号和逗号
                content = content.replace("；", ";").replace("，", ";").replace(",", ";")
                # 2. 按行和分号分割
                raw_lines = re.split(r'[;\n]', content)
                # 3. 清理和重建
                final_emails = [line.strip() for line in raw_lines if line.strip()]
                
                result = "; ".join(final_emails)
                
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, result)
            else:
                entry_widget.delete(0, tk.END)
            
            dialog.destroy()
            
        ttk.Button(btn_frame, text="✅ 确认更新", command=save, style='Accent.TButton').pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="🗑️ 清空内容", command=lambda: text_area.delete(1.0, tk.END)).pack(side=tk.LEFT, padx=5)

    def show_placeholder_menu(self, entry_widget, is_recipient=False):
        """显示占位符菜单（用于主题/收件人等）"""
        menu = tk.Menu(self.root, tearoff=0)
        
        # 收件人特有选项：手动添加邮箱
        if is_recipient:
            menu.add_command(label="✏️ 批量编辑/导入邮箱...", 
                            command=lambda: self.add_manual_email_dialog(entry_widget))
            menu.add_separator()
        
        # 常用占位符
        menu.add_command(label="📅 {DATE} - 当前日期", 
                        command=lambda: self.insert_placeholder(entry_widget, "{DATE}"))
        menu.add_command(label="⏰ {TIME} - 当前时间", 
                        command=lambda: self.insert_placeholder(entry_widget, "{TIME}"))
        menu.add_command(label="📆 {DATETIME} - 日期时间", 
                        command=lambda: self.insert_placeholder(entry_widget, "{DATETIME}"))
        
        # 特殊功能占位符 (仅正文可用，但也允许在其他地方插入，由用户决定)
        menu.add_separator()
        menu.add_command(label="🖼 {SMART_IMAGE} - 智能截图 (自动截取Excel范围)", 
                        command=lambda: self.insert_placeholder(entry_widget, "{SMART_IMAGE}"))
        menu.add_command(label="📊 {EXCEL_DATA} - Excel数据表格", 
                        command=lambda: self.insert_placeholder(entry_widget, "{EXCEL_DATA}"))
        menu.add_command(label="📷 {EXCEL_IMAGE} - Excel截图 (静态)", 
                        command=lambda: self.insert_placeholder(entry_widget, "{EXCEL_IMAGE}"))
        
        # 自定义占位符
        if hasattr(self, 'custom_placeholders') and self.custom_placeholders:
            menu.add_separator()
            menu.add_command(label="自定义占位符:", state="disabled")
            for name, value in self.custom_placeholders.items():
                display_val = str(value)
                if len(display_val) > 15:
                    display_val = display_val[:12] + "..."
                menu.add_command(label=f"  📌 {{{name}}} ({display_val})", 
                               command=lambda n=name: self.insert_placeholder(entry_widget, f"{{{n}}}"))
        
        # 签名占位符 (新增)
        if hasattr(self, 'email_signatures') and self.email_signatures:
            menu.add_separator()
            menu.add_command(label="签名占位符:", state="disabled")
            for name in self.email_signatures.keys():
                menu.add_command(label=f"  ✍️ {{{name}}}", 
                               command=lambda n=name: self.insert_placeholder(entry_widget, f"{{{n}}}"))

        menu.add_separator()
        
        # Excel列占位符
        if self.current_excel_data and len(self.current_excel_data) > 0:
            headers = self.current_excel_data[0]
            menu.add_command(label="Excel 列占位符:", state="disabled")
            for header in headers[:10]:  # 最多显示10个
                menu.add_command(label=f"  📊 {{{header}}}", 
                               command=lambda h=header: self.insert_placeholder(entry_widget, f"{{{h}}}"))
            if len(headers) > 10:
                menu.add_command(label=f"  ... 还有 {len(headers)-10} 个", state="disabled")
        else:
            menu.add_command(label="(先读取Excel数据)", state="disabled")
        
        # 显示菜单
        try:
            menu.post(entry_widget.winfo_rootx(), entry_widget.winfo_rooty() + entry_widget.winfo_height())
        finally:
            menu.grab_release()
    
    def edit_body_in_window(self):
        """在弹出窗口中编辑正文"""
        # 创建弹出窗口
        body_window = tk.Toplevel(self.root)
        body_window.title("编辑邮件正文")
        body_window.geometry("900x600")
        body_window.transient(self.root)
        
        # 工具栏
        toolbar = ttk.Frame(body_window, padding="5")
        toolbar.pack(fill=tk.X, side=tk.TOP)
        
        ttk.Label(toolbar, text="编辑正文内容:", font=('TkDefaultFont', 10, 'bold')).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="📋 插入占位符", command=lambda: self.show_body_placeholder_menu_popup(body_text)).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="✍️ 插入签名", command=lambda: self.insert_signature_menu_popup(body_text)).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="🔍 拼写检查", command=lambda: self.check_spelling_popup(body_text)).pack(side=tk.LEFT, padx=2)
        
        # 正文编辑区
        text_frame = ttk.Frame(body_window, padding="5")
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        body_text = scrolledtext.ScrolledText(text_frame, width=80, height=25, wrap=tk.WORD, font=('TkDefaultFont', 10))
        body_text.pack(fill=tk.BOTH, expand=True)
        
        # 加载当前正文内容
        current_body = self.body_text.get(1.0, tk.END)
        body_text.insert(1.0, current_body)
        
        # 按钮区
        button_frame = ttk.Frame(body_window, padding="10")
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        def save_and_close():
            # 保存内容到隐藏的body_text
            content = body_text.get(1.0, tk.END).strip()
            self.body_text.delete(1.0, tk.END)
            self.body_text.insert(1.0, content)
            
            body_window.destroy()
        
        ttk.Button(button_frame, text="✓ 保存并关闭", command=save_and_close, width=15).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="✗ 取消", command=body_window.destroy, width=15).pack(side=tk.RIGHT, padx=5)
        
        # 聚焦到文本框
        body_text.focus_set()
    
    def show_body_placeholder_menu_popup(self, text_widget):
        """在弹窗中显示正文占位符菜单"""
        menu = tk.Menu(self.root, tearoff=0)
        
        # 常用占位符
        menu.add_command(label="📅 {DATE} - 当前日期", 
                        command=lambda: text_widget.insert(tk.INSERT, "{DATE}"))
        menu.add_command(label="⏰ {TIME} - 当前时间", 
                        command=lambda: text_widget.insert(tk.INSERT, "{TIME}"))
        menu.add_command(label="📆 {DATETIME} - 日期时间", 
                        command=lambda: text_widget.insert(tk.INSERT, "{DATETIME}"))
        
        menu.add_separator()
        menu.add_command(label="📊 {EXCEL_DATA} - 插入完整表格", 
                        command=lambda: text_widget.insert(tk.INSERT, "{EXCEL_DATA}"))
        menu.add_command(label="🖼️ {EXCEL_IMAGE} - 插入Excel截图", 
                        command=lambda: text_widget.insert(tk.INSERT, "{EXCEL_IMAGE}"))
        
        menu.add_separator()
        
        # Excel列占位符
        if self.current_excel_data and len(self.current_excel_data) > 0:
            headers = self.current_excel_data[0]
            menu.add_command(label="Excel 列占位符 (用于批量):", state="disabled")
            for header in headers[:15]:
                menu.add_command(label=f"  📊 {{{header}}}", 
                               command=lambda h=header: text_widget.insert(tk.INSERT, f"{{{h}}}"))
            if len(headers) > 15:
                menu.add_command(label=f"  ... 还有 {len(headers)-15} 个", state="disabled")
        else:
            menu.add_command(label="(先读取Excel数据查看列占位符)", state="disabled")
        
        menu.add_separator()
        
        # 自定义占位符
        if self.custom_placeholders:
            menu.add_command(label="自定义占位符:", state="disabled")
            for name, value in list(self.custom_placeholders.items())[:10]:
                display_value = str(value)[:20] + "..." if len(str(value)) > 20 else str(value)
                menu.add_command(label=f"  {{{name}}} = {display_value}", 
                               command=lambda n=name: text_widget.insert(tk.INSERT, f"{{{n}}}"))
        
        # 显示菜单
        try:
            menu.post(text_widget.winfo_rootx() + 10, text_widget.winfo_rooty() + 30)
        finally:
            menu.grab_release()
    
    def insert_signature_menu_popup(self, text_widget):
        """在弹窗中插入签名"""
        menu = tk.Menu(self.root, tearoff=0)
        
        if not self.email_signatures:
            menu.add_command(label="（暂无签名，请先添加）", state="disabled")
        else:
            for name, content in self.email_signatures.items():
                # 简单的预览处理
                preview = content.strip().replace('\n', ' ')
                if len(preview) > 20:
                    preview = preview[:20] + "..."
                
                menu.add_command(label=f"✍️ {name} ({preview})", 
                               command=lambda c=content: text_widget.insert(tk.INSERT, "\n\n" + c))
        
        menu.add_separator()
        menu.add_command(label="⚙️ 管理签名...", command=self.manage_signatures)
        
        try:
            # 获取按钮位置（如果可能）或者鼠标位置
            x = self.root.winfo_pointerx()
            y = self.root.winfo_pointery()
            menu.post(x, y)
        finally:
            menu.grab_release()
    
    def check_spelling_popup(self, text_widget):
        """在弹窗中检查拼写"""
        content = text_widget.get(1.0, tk.END).strip()
        if not content:
            messagebox.showinfo("提示", "正文为空")
            return
        
        # 检查拼写
        errors = []
        for wrong, correct in self.spell_check_dict.items():
            if wrong in content:
                errors.append((wrong, correct))
        
        if not errors:
            messagebox.showinfo("拼写检查", "✓ 未发现常见拼写错误")
            return
        
        # 显示错误并提供修正
        msg = "发现以下可能的拼写错误:\n\n"
        for wrong, correct in errors:
            msg += f"  • '{wrong}' → '{correct}'\n"
        msg += f"\n是否自动修正这些错误？"
        
        if messagebox.askyesno("拼写检查", msg):
            new_content = content
            for wrong, correct in errors:
                new_content = new_content.replace(wrong, correct)
            text_widget.delete(1.0, tk.END)
            text_widget.insert(1.0, new_content)
            messagebox.showinfo("完成", f"已修正 {len(errors)} 个拼写错误")
    
    def show_body_placeholder_menu(self):
        """显示正文占位符菜单（旧方法，保留兼容性）"""
        menu = tk.Menu(self.root, tearoff=0)
        
        # 常用占位符
        menu.add_command(label="📅 {DATE} - 当前日期", 
                        command=lambda: self.insert_text_placeholder("{DATE}"))
        menu.add_command(label="⏰ {TIME} - 当前时间", 
                        command=lambda: self.insert_text_placeholder("{TIME}"))
        menu.add_command(label="📆 {DATETIME} - 日期时间", 
                        command=lambda: self.insert_text_placeholder("{DATETIME}"))
        
        menu.add_separator()
        menu.add_command(label="📊 {EXCEL_DATA} - 插入完整表格", 
                        command=lambda: self.insert_text_placeholder("{EXCEL_DATA}"))
        menu.add_command(label="🖼️ {EXCEL_IMAGE} - 插入Excel截图", 
                        command=lambda: self.insert_text_placeholder("{EXCEL_IMAGE}"))
        
        menu.add_separator()
        
        # Excel列占位符
        if self.current_excel_data and len(self.current_excel_data) > 0:
            headers = self.current_excel_data[0]
            menu.add_command(label="Excel 列占位符 (用于批量):", state="disabled")
            for header in headers[:15]:
                menu.add_command(label=f"  📊 {{{header}}}", 
                               command=lambda h=header: self.insert_text_placeholder(f"{{{h}}}"))
            if len(headers) > 15:
                menu.add_command(label=f"  ... 还有 {len(headers)-15} 个", state="disabled")
        else:
            menu.add_command(label="(先读取Excel数据查看列占位符)", state="disabled")
        
        menu.add_separator()
        
        # 自定义占位符
        menu.add_command(label="✏️ 自定义占位符...", 
                        command=self.create_custom_placeholder)
        
        # 显示菜单
        try:
            menu.post(self.body_text.winfo_rootx(), self.body_text.winfo_rooty() + 30)
        finally:
            menu.grab_release()
    
    def insert_placeholder(self, entry_widget, placeholder):
        """在Widget中插入占位符"""
        if isinstance(entry_widget, (ttk.Entry, tk.Entry)):
            # 获取当前光标位置
            try:
                current_pos = entry_widget.index(tk.INSERT)
                entry_widget.insert(current_pos, placeholder)
                entry_widget.focus()
            except:
                pass
        elif isinstance(entry_widget, (tk.Text, scrolledtext.ScrolledText)):
            try:
                entry_widget.insert(tk.INSERT, placeholder)
                entry_widget.focus()
            except:
                pass
            
            # 移动光标到占位符后
            # entry_widget.icursor(current_pos + len(placeholder))
            # entry_widget.focus()
    
    def insert_text_placeholder(self, placeholder):
        """在Text控件中插入占位符"""
        # 获取当前光标位置
        cursor_pos = self.body_text.index(tk.INSERT)
        
        # 插入占位符
        self.body_text.insert(cursor_pos, placeholder)
        
        # 移动光标到占位符后
        self.body_text.mark_set(tk.INSERT, f"{cursor_pos}+{len(placeholder)}c")
        self.body_text.focus()
    
    def create_custom_placeholder(self):
        """创建自定义占位符"""
        dialog = tk.Toplevel(self.root)
        dialog.title("创建自定义占位符")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="输入占位符名称（不含花括号）：", 
                 font=('TkDefaultFont', 10)).pack(pady=10)
        
        name_var = tk.StringVar()
        entry = ttk.Entry(dialog, textvariable=name_var, width=30)
        entry.pack(pady=5)
        entry.focus()
        
        ttk.Label(dialog, text="例如: 客户名称, 订单号, 金额 等", 
                 font=('TkDefaultFont', 8), foreground='gray').pack()
        
        def insert_custom():
            name = name_var.get().strip()
            if name:
                placeholder = "{" + name + "}"
                self.insert_text_placeholder(placeholder)
                dialog.destroy()
            else:
                messagebox.showwarning("警告", "请输入占位符名称")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="插入", command=insert_custom).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        # 绑定回车键
        entry.bind('<Return>', lambda e: insert_custom())
    
    def add_custom_placeholder(self):
        """添加自定义占位符"""
        name = self.placeholder_name_var.get().strip()
        value = self.placeholder_value_var.get().strip()
        
        if not name:
            messagebox.showwarning("警告", "请输入占位符名称")
            return
        
        if not value:
            messagebox.showwarning("警告", "请输入占位符的值")
            return
        
        # 保存占位符
        self.custom_placeholders[name] = value
        
        # 清空输入
        self.placeholder_name_var.set("")
        self.placeholder_value_var.set("")
        
        messagebox.showinfo("成功", f"已添加占位符: {{{name}}} = {value}\n\n在主题或正文中使用 {{{name}}} 即可自动替换")
        self.status_var.set(f"已添加占位符: {{{name}}}")
    
    def remove_custom_placeholder(self):
        """从管理窗口删除选中的占位符"""
        if hasattr(self, 'placeholder_listbox_manage'):
            selection = self.placeholder_listbox_manage.curselection()
            if not selection:
                messagebox.showwarning("警告", "请先选择要删除的占位符")
                return
            
            index = selection[0]
            keys = list(self.custom_placeholders.keys())
            if index < len(keys):
                key = keys[index]
                del self.custom_placeholders[key]
                self.update_placeholder_listbox()
                messagebox.showinfo("成功", f"已删除占位符: {{{key}}}")
    
    def clear_custom_placeholders(self):
        """清空所有自定义占位符"""
        if not self.custom_placeholders:
            messagebox.showinfo("提示", "没有自定义占位符")
            return
        
        if messagebox.askyesno("确认", f"确定要清空所有 {len(self.custom_placeholders)} 个自定义占位符吗？"):
            self.custom_placeholders.clear()
            if hasattr(self, 'placeholder_listbox_manage'):
                self.update_placeholder_listbox()
            messagebox.showinfo("成功", "已清空所有自定义占位符")
    
    def manage_placeholders(self):
        """打开占位符管理窗口 - 增强版"""
        dialog = tk.Toplevel(self.root)
        dialog.title("占位符管理")
        dialog.geometry("600x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 说明文本
        ttk.Label(dialog, text="管理自定义占位符 (可在主题或正文中使用 {占位符名})", 
                 font=('TkDefaultFont', 10, 'bold')).pack(pady=(10, 5), padx=10)
        
        # 列表区域
        list_frame = ttk.Frame(dialog)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Treeview
        columns = ("name", "value")
        self.ph_tree = ttk.Treeview(list_frame, columns=columns, show="headings", selectmode="extended")
        self.ph_tree.heading("name", text="占位符名称")
        self.ph_tree.heading("value", text="替换值")
        self.ph_tree.column("name", width=150)
        self.ph_tree.column("value", width=300)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.ph_tree.yview)
        self.ph_tree.configure(yscrollcommand=scrollbar.set)
        
        self.ph_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.update_placeholder_tree()
        
        # 编辑区域
        edit_frame = ttk.LabelFrame(dialog, text="编辑/添加", padding="10")
        edit_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(edit_frame, text="名称:").grid(row=0, column=0, padx=5)
        name_var = tk.StringVar()
        name_entry = ttk.Entry(edit_frame, textvariable=name_var, width=20)
        name_entry.grid(row=0, column=1, padx=5)
        
        ttk.Label(edit_frame, text="值:").grid(row=0, column=2, padx=5)
        value_var = tk.StringVar()
        value_entry = ttk.Entry(edit_frame, textvariable=value_var, width=30)
        value_entry.grid(row=0, column=3, padx=5)
        
        def on_select(event):
            selected_items = self.ph_tree.selection()
            if selected_items:
                item = self.ph_tree.item(selected_items[0])
                name_var.set(item['values'][0])
                value_var.set(item['values'][1])
        
        self.ph_tree.bind("<<TreeviewSelect>>", on_select)
        
        def save_item():
            name = name_var.get().strip()
            value = value_var.get().strip()
            if not name:
                messagebox.showwarning("警告", "请输入占位符名称")
                return
            # 移除可能的大括号
            name = name.replace("{", "").replace("}", "")
            self.custom_placeholders[name] = value
            self.update_placeholder_tree()
            name_var.set("")
            value_var.set("")
            
        def delete_item():
            selected_items = self.ph_tree.selection()
            if not selected_items:
                return
            if messagebox.askyesno("确认", "确定删除选中的占位符吗？"):
                for item_id in selected_items:
                    item = self.ph_tree.item(item_id)
                    name = item['values'][0]
                    if name in self.custom_placeholders:
                        del self.custom_placeholders[name]
                self.update_placeholder_tree()
        
        def batch_edit():
            """批量编辑窗口"""
            batch_win = tk.Toplevel(dialog)
            batch_win.title("批量编辑占位符")
            batch_win.geometry("500x400")
            
            ttk.Label(batch_win, text="请输入JSON格式 (例如: {\"公司\": \"ABC\", \"年份\": \"2025\"})").pack(pady=5)
            text_area = scrolledtext.ScrolledText(batch_win, width=60, height=15)
            text_area.pack(padx=10, pady=5)
            
            # 预填当前数据
            import json
            text_area.insert(1.0, json.dumps(self.custom_placeholders, ensure_ascii=False, indent=2))
            
            def apply_batch():
                try:
                    content = text_area.get(1.0, tk.END).strip()
                    new_data = json.loads(content)
                    if isinstance(new_data, dict):
                        self.custom_placeholders.update(new_data)
                        self.update_placeholder_tree()
                        batch_win.destroy()
                        messagebox.showinfo("成功", "批量更新成功")
                    else:
                        messagebox.showerror("错误", "必须是JSON对象格式")
                except Exception as e:
                    messagebox.showerror("错误", f"解析JSON失败: {e}")
            
            ttk.Button(batch_win, text="应用更改", command=apply_batch).pack(pady=10)

        btn_frame = ttk.Frame(edit_frame)
        btn_frame.grid(row=1, column=0, columnspan=4, pady=10)
        
        ttk.Button(btn_frame, text="💾 保存/更新", command=save_item).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ 删除选中", command=delete_item).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📝 批量编辑(JSON)", command=batch_edit).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🔄 同步到所有配置", command=lambda: self.sync_placeholders_to_all_configs(dialog)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def sync_placeholders_to_all_configs(self, parent_dialog=None):
        """将当前占位符同步/更新到所有已保存的配置"""
        if not self.configs:
            messagebox.showinfo("提示", "没有已保存的配置", parent=parent_dialog)
            return
        
        config_names = list(self.configs.keys())
        current_ph = self.custom_placeholders.copy()
        
        # 弹出确认对话框，让用户选择同步模式
        sync_dialog = tk.Toplevel(parent_dialog or self.root)
        sync_dialog.title("同步占位符到所有配置")
        sync_dialog.geometry("500x400")
        sync_dialog.transient(parent_dialog or self.root)
        sync_dialog.grab_set()
        
        ttk.Label(sync_dialog, text="将当前占位符同步到已保存的配置", 
                 font=('TkDefaultFont', 10, 'bold')).pack(pady=(10, 5))
        
        # 同步模式
        mode_frame = ttk.LabelFrame(sync_dialog, text="同步模式", padding="10")
        mode_frame.pack(fill=tk.X, padx=10, pady=5)
        
        sync_mode = tk.StringVar(value="merge")
        ttk.Radiobutton(mode_frame, text="合并 - 新增/更新占位符，保留配置中已有的其他占位符", 
                        variable=sync_mode, value="merge").pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(mode_frame, text="覆盖 - 用当前占位符完全替换所有配置的占位符", 
                        variable=sync_mode, value="replace").pack(anchor=tk.W, pady=2)
        
        # 显示当前占位符
        ph_frame = ttk.LabelFrame(sync_dialog, text=f"当前占位符 ({len(current_ph)} 个)", padding="5")
        ph_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ph_text = ""
        for name, value in current_ph.items():
            display_val = str(value)
            if len(display_val) > 30:
                display_val = display_val[:27] + "..."
            ph_text += f"  {{{name}}} = {display_val}\n"
        if not ph_text:
            ph_text = "  (无占位符)"
        
        ph_label = tk.Text(ph_frame, height=min(6, max(2, len(current_ph))), wrap=tk.WORD, font=('TkDefaultFont', 9))
        ph_label.insert(1.0, ph_text.strip())
        ph_label.config(state='disabled')
        ph_label.pack(fill=tk.X)
        
        # 选择要同步的配置
        config_frame = ttk.LabelFrame(sync_dialog, text="选择要同步的配置", padding="5")
        config_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 全选/取消
        select_all_var = tk.BooleanVar(value=True)
        config_vars = {}
        
        def toggle_all():
            val = select_all_var.get()
            for v in config_vars.values():
                v.set(val)
        
        ttk.Checkbutton(config_frame, text="全选/取消全选", variable=select_all_var, 
                       command=toggle_all).pack(anchor=tk.W)
        
        config_list_frame = ttk.Frame(config_frame)
        config_list_frame.pack(fill=tk.BOTH, expand=True)
        
        config_canvas = tk.Canvas(config_list_frame, height=100)
        config_scrollbar = ttk.Scrollbar(config_list_frame, orient="vertical", command=config_canvas.yview)
        config_inner = ttk.Frame(config_canvas)
        
        config_inner.bind("<Configure>", lambda e: config_canvas.configure(scrollregion=config_canvas.bbox("all")))
        config_canvas.create_window((0, 0), window=config_inner, anchor="nw")
        config_canvas.configure(yscrollcommand=config_scrollbar.set)
        
        config_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        config_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        for name in config_names:
            var = tk.BooleanVar(value=True)
            config_vars[name] = var
            existing_ph_count = len(self.configs[name].get("custom_placeholders", {}))
            ttk.Checkbutton(config_inner, text=f"{name} (现有 {existing_ph_count} 个占位符)", 
                          variable=var).pack(anchor=tk.W, padx=10)
        
        # 执行同步
        def do_sync():
            selected_configs = [name for name, var in config_vars.items() if var.get()]
            if not selected_configs:
                messagebox.showwarning("警告", "请至少选择一个配置", parent=sync_dialog)
                return
            
            mode = sync_mode.get()
            updated_count = 0
            
            for config_name in selected_configs:
                if config_name in self.configs:
                    if mode == "replace":
                        self.configs[config_name]["custom_placeholders"] = current_ph.copy()
                    else:  # merge
                        existing = self.configs[config_name].get("custom_placeholders", {})
                        existing.update(current_ph)
                        self.configs[config_name]["custom_placeholders"] = existing
                    updated_count += 1
            
            if self.save_configs():
                sync_dialog.destroy()
                messagebox.showinfo("成功", 
                    f"已将 {len(current_ph)} 个占位符{'覆盖' if mode == 'replace' else '合并'}同步到 {updated_count} 个配置。",
                    parent=parent_dialog or self.root)
                self.status_var.set(f"占位符已同步到 {updated_count} 个配置")
        
        # 按钮
        btn_frame = ttk.Frame(sync_dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        ttk.Button(btn_frame, text="✅ 执行同步", command=do_sync).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=sync_dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def update_placeholder_tree(self):
        """更新占位符Treeview"""
        if hasattr(self, 'ph_tree'):
            for item in self.ph_tree.get_children():
                self.ph_tree.delete(item)
            
            for name, value in self.custom_placeholders.items():
                self.ph_tree.insert('', tk.END, values=(name, value))
    
    def load_signatures(self):
        """加载邮件签名"""
        signature_file = os.path.join(self._base_dir, "email_signatures.json")
        if os.path.exists(signature_file):
            try:
                with open(signature_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"加载签名文件失败: {e}")
        # 默认签名
        return {
            "默认签名": "\n\n--\n此致\n敬礼！",
            "正式签名": "\n\n───────────────\n{公司名}\n{联系人}\n电话: {电话}\n邮箱: {邮箱}",
            "简洁签名": "\n\n谢谢！\nBest regards"
        }
    
    def save_signatures(self):
        """保存邮件签名"""
        signature_file = os.path.join(self._base_dir, "email_signatures.json")
        try:
            with open(signature_file, 'w', encoding='utf-8') as f:
                json.dump(self.email_signatures, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror("错误", f"保存签名失败: {str(e)}")
            return False

    def load_content_templates(self):
        """加载内容模板"""
        template_file = os.path.join(self._base_dir, "content_templates.json")
        if os.path.exists(template_file):
            try:
                with open(template_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"加载内容模板失败: {e}")
        # 默认模板
        return {
            "默认模板": {
                "subject": "关于 {DATE} 的工作汇报",
                "body": "您好，\n\n这是 {DATE} 的工作汇报，请查收附件。\n\n谢谢！"
            },
            "会议通知": {
                "subject": "会议通知：{主题}",
                "body": "各位同事：\n\n我们将于 {TIME} 举行会议，请准时参加。\n\n地点：会议室\n议题：{主题}\n\n收到请回复。"
            }
        }

    def save_content_templates(self):
        """保存内容模板"""
        template_file = os.path.join(self._base_dir, "content_templates.json")
        try:
            with open(template_file, 'w', encoding='utf-8') as f:
                json.dump(self.content_templates, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror("错误", f"保存模板失败: {str(e)}")
            return False

    def add_format_toolbar(self, parent, text_widget):
        """添加格式化工具栏 (加粗、斜体等)"""
        f_toolbar = ttk.Frame(parent)
        f_toolbar.pack(fill=tk.X, pady=2)
        
        def insert_tag(open_tag, close_tag):
            try:
                # 获取选区
                sel_start = text_widget.index("sel.first")
                sel_end = text_widget.index("sel.last")
                selection = text_widget.get(sel_start, sel_end)
                
                # 插入标签
                text_widget.delete(sel_start, sel_end)
                text_widget.insert(sel_start, f"{open_tag}{selection}{close_tag}")
            except tk.TclError:
                # 没有选区，直接在光标处插入
                text_widget.insert(tk.INSERT, f"{open_tag}{close_tag}")
                # 将光标移动到标签中间
                text_widget.mark_set(tk.INSERT, f"{tk.INSERT}-{len(close_tag)}c")
        
        ttk.Button(f_toolbar, text="B", width=3, command=lambda: insert_tag("<b>", "</b>")).pack(side=tk.LEFT, padx=1)
        ttk.Button(f_toolbar, text="I", width=3, command=lambda: insert_tag("<i>", "</i>")).pack(side=tk.LEFT, padx=1)
        ttk.Button(f_toolbar, text="U", width=3, command=lambda: insert_tag("<u>", "</u>")).pack(side=tk.LEFT, padx=1)
        ttk.Label(f_toolbar, text="|").pack(side=tk.LEFT, padx=5)
        
        ttk.Button(f_toolbar, text="H1", width=3, command=lambda: insert_tag("<h1>", "</h1>")).pack(side=tk.LEFT, padx=1)
        ttk.Button(f_toolbar, text="H2", width=3, command=lambda: insert_tag("<h2>", "</h2>")).pack(side=tk.LEFT, padx=1)
        ttk.Label(f_toolbar, text="|").pack(side=tk.LEFT, padx=5)
        
        ttk.Button(f_toolbar, text="红字", width=5, command=lambda: insert_tag('<span style="color:red">', '</span>')).pack(side=tk.LEFT, padx=1)
        ttk.Button(f_toolbar, text="换行", width=5, command=lambda: text_widget.insert(tk.INSERT, "<br>")).pack(side=tk.LEFT, padx=1)

    def manage_content_templates(self):
        """管理内容模板"""
        dialog = tk.Toplevel(self.root)
        dialog.title("草稿内容模板管理")
        dialog.geometry("900x700")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 左右布局
        paned = ttk.PanedWindow(dialog, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 左侧列表
        left_frame = ttk.LabelFrame(paned, text="模板列表", padding="5")
        paned.add(left_frame, weight=1)
        
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        template_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, font=('TkDefaultFont', 10))
        template_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=template_listbox.yview)
        
        # 右侧编辑
        right_frame = ttk.LabelFrame(paned, text="模板内容", padding="5")
        paned.add(right_frame, weight=2)
        
        ttk.Label(right_frame, text="模板名称:").pack(anchor=tk.W)
        name_var = tk.StringVar()
        ttk.Entry(right_frame, textvariable=name_var).pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(right_frame, text="主题:").pack(anchor=tk.W)
        subject_var = tk.StringVar()
        ttk.Entry(right_frame, textvariable=subject_var).pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(right_frame, text="正文:").pack(anchor=tk.W)
        
        # 增加正文工具栏容器
        toolbars_container = ttk.Frame(right_frame)
        toolbars_container.pack(fill=tk.X, pady=(0, 2))
        
        # 1. 占位符栏
        ph_toolbar = ttk.Frame(toolbars_container)
        ph_toolbar.pack(fill=tk.X)
        
        # 2. 格式化栏
        fmt_toolbar = ttk.Frame(toolbars_container)
        fmt_toolbar.pack(fill=tk.X)

        body_text = scrolledtext.ScrolledText(right_frame, width=40, height=15, wrap=tk.WORD)
        body_text.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        # 填充工具栏
        def insert_ph_menu():
             self.show_placeholder_menu(body_text)

        ttk.Button(ph_toolbar, text="插入占位符", command=insert_ph_menu).pack(side=tk.LEFT, padx=0)
        ttk.Label(ph_toolbar, text="(支持 {签名名}, {自定义占位符} 等)", font=('TkDefaultFont', 8), foreground='gray').pack(side=tk.LEFT, padx=5)

        # 添加格式化工具栏
        self.add_format_toolbar(fmt_toolbar, body_text)


        
        # 按钮区域
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def update_list():
            template_listbox.delete(0, tk.END)
            for name in sorted(self.content_templates.keys()):
                template_listbox.insert(tk.END, name)
        
        def on_select(event):
            selection = template_listbox.curselection()
            if selection:
                name = template_listbox.get(selection[0])
                data = self.content_templates.get(name, {})
                name_var.set(name)
                subject_var.set(data.get("subject", ""))
                body_text.delete(1.0, tk.END)
                body_text.insert(1.0, data.get("body", ""))
        
        template_listbox.bind('<<ListboxSelect>>', on_select)
        
        def save_template():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("警告", "请输入模板名称")
                return
            
            subject = subject_var.get()
            body = body_text.get(1.0, tk.END).strip() # 保留换行，但去除首尾空白
            
            self.content_templates[name] = {
                "subject": subject,
                "body": body
            }
            if self.save_content_templates():
                update_list()
                messagebox.showinfo("成功", "模板已保存")
        
        def delete_template():
            selection = template_listbox.curselection()
            if not selection:
                return
            name = template_listbox.get(selection[0])
            if messagebox.askyesno("确认", f"确定删除模板 '{name}' 吗？"):
                del self.content_templates[name]
                self.save_content_templates()
                update_list()
                # 清空右侧
                name_var.set("")
                subject_var.set("")
                body_text.delete(1.0, tk.END)
        
        def apply_template():
            selection = template_listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请先选择一个模板")
                return
            name = template_listbox.get(selection[0])
            data = self.content_templates.get(name, {})
            
            # 应用到主界面
            if messagebox.askyesno("确认", f"确定要应用模板 '{name}' 吗？\n这将覆盖当前的主题和正文。"):
                self.subject_var.set(data.get("subject", ""))
                self.body_text.delete(1.0, tk.END)
                self.body_text.insert(1.0, data.get("body", ""))
                dialog.destroy()
                self.status_var.set(f"已应用模板: {name}")

        def save_current_as_new():
            # 获取主界面当前内容
            current_subject = self.subject_var.get()
            current_body = self.body_text.get(1.0, tk.END).strip()
            
            name_var.set("新模板")
            subject_var.set(current_subject)
            body_text.delete(1.0, tk.END)
            body_text.insert(1.0, current_body)
            messagebox.showinfo("提示", "已获取当前内容，请修改名称后保存")

        # 左侧按钮
        left_btn_frame = ttk.Frame(left_frame)
        left_btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(left_btn_frame, text="应用选中模板", command=apply_template).pack(fill=tk.X)
        
        # 底部按钮
        def clear_fields():
            name_var.set("")
            subject_var.set("")
            body_text.delete(1.0, tk.END)
            # 清除选中
            template_listbox.selection_clear(0, tk.END)

        ttk.Button(btn_frame, text="✨ 清空/新建", command=clear_fields).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="从当前内容新建", command=save_current_as_new).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="保存/更新模板", command=save_template).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="删除模板", command=delete_template).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)
        
        update_list()
    
    def insert_signature_menu(self):
        """显示插入签名菜单"""
        menu = tk.Menu(self.root, tearoff=0)
        
        if not self.email_signatures:
            menu.add_command(label="（暂无签名，请先添加）", state="disabled")
        else:
            for sig_name in self.email_signatures.keys():
                menu.add_command(
                    label=f"✍️ {sig_name}",
                    command=lambda name=sig_name: self.insert_signature(name)
                )
        
        menu.add_separator()
        menu.add_command(label="⚙️ 管理签名...", command=self.manage_signatures)
        
        try:
            menu.post(self.body_text.winfo_rootx(), self.body_text.winfo_rooty() + 30)
        finally:
            menu.grab_release()
    
    def insert_signature(self, signature_name):
        """插入签名到正文"""
        if signature_name in self.email_signatures:
            signature = self.email_signatures[signature_name]
            current_text = self.body_text.get(1.0, tk.END).rstrip()
            self.body_text.delete(1.0, tk.END)
            self.body_text.insert(1.0, current_text + signature)
            self.status_var.set(f"已插入签名: {signature_name}")
    
    def manage_signatures(self):
        """管理邮件签名"""
        dialog = tk.Toplevel(self.root)
        dialog.title("邮件签名管理")
        dialog.geometry("600x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 说明
        ttk.Label(dialog, text="管理您的邮件签名模板，可以在正文中快速插入", 
                 font=('TkDefaultFont', 9), foreground='blue').pack(pady=10, padx=10)
        
        # 签名列表
        list_frame = ttk.LabelFrame(dialog, text="已保存的签名", padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 列表框
        list_container = ttk.Frame(list_frame)
        list_container.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.signature_listbox = tk.Listbox(list_container, yscrollcommand=scrollbar.set, font=('TkDefaultFont', 10))
        self.signature_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.signature_listbox.yview)
        
        self.signature_listbox.bind('<<ListboxSelect>>', self.on_signature_select)
        self.update_signature_listbox()
        
        # 编辑区域
        edit_frame = ttk.LabelFrame(dialog, text="签名内容", padding="10")
        edit_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        ttk.Label(edit_frame, text="签名名称:").pack(anchor=tk.W)
        self.signature_name_entry = ttk.Entry(edit_frame, width=40)
        self.signature_name_entry.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(edit_frame, text="签名内容: (支持占位符，如 {公司名})").pack(anchor=tk.W)
        self.signature_text = scrolledtext.ScrolledText(edit_frame, width=60, height=6, wrap=tk.WORD)
        self.signature_text.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        # 按钮区域
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame, text="➕ 新建", command=self.new_signature).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="💾 保存", command=self.save_signature).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ 删除", command=self.delete_signature).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def update_signature_listbox(self):
        """更新签名列表"""
        if hasattr(self, 'signature_listbox'):
            self.signature_listbox.delete(0, tk.END)
            for name in self.email_signatures.keys():
                self.signature_listbox.insert(tk.END, name)
    
    def on_signature_select(self, event):
        """选择签名时显示内容"""
        selection = self.signature_listbox.curselection()
        if selection:
            index = selection[0]
            sig_name = self.signature_listbox.get(index)
            if sig_name in self.email_signatures:
                self.signature_name_entry.delete(0, tk.END)
                self.signature_name_entry.insert(0, sig_name)
                self.signature_text.delete(1.0, tk.END)
                self.signature_text.insert(1.0, self.email_signatures[sig_name])
    
    def new_signature(self):
        """新建签名"""
        self.signature_name_entry.delete(0, tk.END)
        self.signature_text.delete(1.0, tk.END)
        self.signature_name_entry.focus()
    
    def save_signature(self):
        """保存/更新签名"""
        sig_name = self.signature_name_entry.get().strip()
        sig_content = self.signature_text.get(1.0, tk.END).rstrip()
        
        if not sig_name:
            messagebox.showwarning("警告", "请输入签名名称")
            return
        
        if not sig_content:
            messagebox.showwarning("警告", "请输入签名内容")
            return
        
        self.email_signatures[sig_name] = sig_content
        if self.save_signatures():
            self.update_signature_listbox()
            messagebox.showinfo("成功", f"签名 '{sig_name}' 已保存")
    
    def delete_signature(self):
        """删除签名"""
        sig_name = self.signature_name_entry.get().strip()
        
        if not sig_name:
            messagebox.showwarning("警告", "请先选择要删除的签名")
            return
        
        if sig_name in self.email_signatures:
            if messagebox.askyesno("确认", f"确定要删除签名 '{sig_name}' 吗？"):
                del self.email_signatures[sig_name]
                if self.save_signatures():
                    self.update_signature_listbox()
                    self.new_signature()
                    messagebox.showinfo("成功", "签名已删除")
    
    def check_spelling(self):
        """拼写检查"""
        subject = self.subject_var.get()
        body = self.body_text.get(1.0, tk.END)
        
        errors_found = []
        
        # 检查主题
        for wrong, correct in self.spell_check_dict.items():
            if wrong in subject:
                errors_found.append(f"主题中: '{wrong}' 应为 '{correct}'")
        
        # 检查正文
        for wrong, correct in self.spell_check_dict.items():
            if wrong in body:
                errors_found.append(f"正文中: '{wrong}' 应为 '{correct}'")
        
        # 检查常见格式问题
        if re.search(r'\s{2,}', subject):
            errors_found.append("主题中有多余的空格")
        
        if re.search(r'[a-zA-Z][，。！？]', body):
            errors_found.append("正文中英文后使用了中文标点")
        
        if re.search(r'[\u4e00-\u9fa5][,\.!?]', body):
            errors_found.append("正文中中文后使用了英文标点")
        
        if errors_found:
            error_msg = "发现以下可能的问题:\n\n" + "\n".join(f"• {e}" for e in errors_found)
            error_msg += "\n\n是否自动修正?"
            
            if messagebox.askyesno("拼写检查", error_msg):
                # 自动修正
                new_subject = subject
                new_body = body
                
                for wrong, correct in self.spell_check_dict.items():
                    new_subject = new_subject.replace(wrong, correct)
                    new_body = new_body.replace(wrong, correct)
                
                self.subject_var.set(new_subject)
                self.body_text.delete(1.0, tk.END)
                self.body_text.insert(1.0, new_body)
                
                messagebox.showinfo("成功", "已自动修正拼写错误")
        else:
            messagebox.showinfo("拼写检查", "未发现明显的拼写错误 ✓")
    
    def schedule_send_dialog(self):
        """定时发送对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("定时发送设置")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="设置邮件的发送时间", font=('TkDefaultFont', 10, 'bold')).pack(pady=10)
        
        # 日期选择
        date_frame = ttk.Frame(dialog)
        date_frame.pack(pady=10, padx=20, fill=tk.X)
        
        ttk.Label(date_frame, text="日期:").pack(side=tk.LEFT, padx=5)
        
        self.schedule_year = tk.StringVar(value=str(datetime.now().year))
        ttk.Spinbox(date_frame, from_=2025, to=2030, textvariable=self.schedule_year, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="年").pack(side=tk.LEFT)
        
        self.schedule_month = tk.StringVar(value=str(datetime.now().month))
        ttk.Spinbox(date_frame, from_=1, to=12, textvariable=self.schedule_month, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="月").pack(side=tk.LEFT)
        
        self.schedule_day = tk.StringVar(value=str(datetime.now().day))
        ttk.Spinbox(date_frame, from_=1, to=31, textvariable=self.schedule_day, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="日").pack(side=tk.LEFT)
        
        # 时间选择
        time_frame = ttk.Frame(dialog)
        time_frame.pack(pady=10, padx=20, fill=tk.X)
        
        ttk.Label(time_frame, text="时间:").pack(side=tk.LEFT, padx=5)
        
        self.schedule_hour = tk.StringVar(value="09")
        ttk.Spinbox(time_frame, from_=0, to=23, textvariable=self.schedule_hour, width=5, format="%02.0f").pack(side=tk.LEFT, padx=2)
        ttk.Label(time_frame, text=":").pack(side=tk.LEFT)
        
        self.schedule_minute = tk.StringVar(value="00")
        ttk.Spinbox(time_frame, from_=0, to=59, textvariable=self.schedule_minute, width=5, format="%02.0f").pack(side=tk.LEFT, padx=2)
        
        # 说明文本
        info_text = """
注意：
• 邮件将先创建为草稿
• 在指定时间自动发送
• 请确保程序在发送时间保持运行
• 发送前会再次确认
        """
        ttk.Label(dialog, text=info_text, font=('TkDefaultFont', 8), foreground='gray', justify=tk.LEFT).pack(pady=10)
        
        # 按钮
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="确定", command=lambda: self.confirm_schedule_send(dialog)).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)
    
    def confirm_schedule_send(self, dialog):
        """确认定时发送"""
        try:
            year = int(self.schedule_year.get())
            month = int(self.schedule_month.get())
            day = int(self.schedule_day.get())
            hour = int(self.schedule_hour.get())
            minute = int(self.schedule_minute.get())
            
            send_time = datetime(year, month, day, hour, minute)
            
            if send_time <= datetime.now():
                messagebox.showerror("错误", "发送时间必须晚于当前时间")
                return
            
            dialog.destroy()
            
            # 创建草稿
            self.create_draft()
            
            # 启动定时发送线程
            time_str = send_time.strftime("%Y-%m-%d %H:%M")
            if messagebox.askyesno("确认", f"草稿已创建\n\n将在 {time_str} 自动发送\n\n请保持程序运行"):
                thread = threading.Thread(target=self.scheduled_send_worker, args=(send_time,), daemon=True)
                thread.start()
                self.status_var.set(f"已设置定时发送: {time_str}")
        
        except Exception as e:
            messagebox.showerror("错误", f"设置失败: {str(e)}")
    
    def scheduled_send_worker(self, send_time):
        """定时发送工作线程"""
        while datetime.now() < send_time:
            time.sleep(60)  # 每分钟检查一次
        
        # 时间到了，尝试发送
        try:
            # 这里需要找到并发送对应的草稿
            # 简化实现：提示用户手动发送
            self.root.after(0, lambda: messagebox.showinfo(
                "定时发送提醒", 
                f"定时发送时间已到！\n\n请在 Outlook 中找到草稿并发送。"
            ))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"发送失败: {str(e)}"))
    
    def show_enhanced_preview(self):
        """显示增强预览（类似Outlook样式）"""
        # 获取邮件信息
        to_list = [x.strip() for x in self.to_entry_var.get().split(';') if x.strip()]
        cc_list = [x.strip() for x in self.cc_entry_var.get().split(';') if x.strip()]
        bcc_list = [x.strip() for x in self.bcc_entry_var.get().split(';') if x.strip()]
        subject = self.subject_var.get()
        body = self.body_text.get(1.0, tk.END).strip()
        
        if not subject:
            messagebox.showwarning("警告", "请填写主题")
            return
        
        # 创建预览窗口
        preview_win = tk.Toplevel(self.root)
        preview_win.title("邮件预览 - Outlook 风格")
        preview_win.geometry("800x600")
        
        # 顶部工具栏
        toolbar = ttk.Frame(preview_win, relief=tk.RAISED, borderwidth=1)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(toolbar, text="📧 邮件预览", font=('TkDefaultFont', 12, 'bold')).pack(side=tk.LEFT, padx=10)
        ttk.Button(toolbar, text="关闭", command=preview_win.destroy).pack(side=tk.RIGHT, padx=5)
        
        # 邮件头部信息
        header_frame = ttk.Frame(preview_win, relief=tk.GROOVE, borderwidth=2)
        header_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 发件人
        from_frame = ttk.Frame(header_frame)
        from_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Label(from_frame, text="发件人:", font=('TkDefaultFont', 9, 'bold'), width=10).pack(side=tk.LEFT)
        ttk.Label(from_frame, text="当前 Outlook 用户", foreground='blue').pack(side=tk.LEFT)
        
        # 收件人
        if to_list:
            to_frame = ttk.Frame(header_frame)
            to_frame.pack(fill=tk.X, padx=10, pady=3)
            ttk.Label(to_frame, text="收件人:", font=('TkDefaultFont', 9, 'bold'), width=10).pack(side=tk.LEFT)
            ttk.Label(to_frame, text="; ".join(to_list), foreground='blue').pack(side=tk.LEFT)
        
        # 抄送
        if cc_list:
            cc_frame = ttk.Frame(header_frame)
            cc_frame.pack(fill=tk.X, padx=10, pady=3)
            ttk.Label(cc_frame, text="抄送:", font=('TkDefaultFont', 9, 'bold'), width=10).pack(side=tk.LEFT)
            ttk.Label(cc_frame, text="; ".join(cc_list), foreground='blue').pack(side=tk.LEFT)
        
        # 主题
        subject_frame = ttk.Frame(header_frame)
        subject_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Label(subject_frame, text="主题:", font=('TkDefaultFont', 9, 'bold'), width=10).pack(side=tk.LEFT)
        ttk.Label(subject_frame, text=subject, font=('TkDefaultFont', 10)).pack(side=tk.LEFT)
        
        # 附件
        if self.attachments:
            attach_frame = ttk.Frame(header_frame)
            attach_frame.pack(fill=tk.X, padx=10, pady=3)
            ttk.Label(attach_frame, text="附件:", font=('TkDefaultFont', 9, 'bold'), width=10).pack(side=tk.LEFT)
            attach_text = f"📎 {len(self.attachments)} 个文件"
            ttk.Label(attach_frame, text=attach_text, foreground='gray').pack(side=tk.LEFT)
        
        # 分隔线
        ttk.Separator(preview_win, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=10, pady=10)
        
        # 邮件正文
        body_frame = ttk.LabelFrame(preview_win, text="正文内容", padding=10)
        body_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 使用 Text 显示正文（支持基本的 HTML 显示）
        body_text = scrolledtext.ScrolledText(body_frame, wrap=tk.WORD, font=('Calibri', 11))
        body_text.pack(fill=tk.BOTH, expand=True)
        
        # 处理占位符
        # 使用统一的变量处理逻辑
        processed_body = self.process_template_variables(body)
        
        # 特殊处理 {SMART_IMAGE} 预览
        smart_img_path = None
        if "{SMART_IMAGE}" in body:
             try:
                 # 获取当前Excel配置 (解析路径变量)
                 excel_path = self.excel_path_var.get()
                 if excel_path:
                     excel_path = self.process_template_variables(excel_path).split(';')[0]
                     if self.smart_match_var.get():
                          excel_path = self.get_latest_file_in_folder(excel_path)
                 
                 sheet_name = self.sheet_combo.get()
                 data_range = self.range_var.get()
                 
                 if excel_path and os.path.exists(excel_path):
                     # 调用逻辑生成图片
                     _, img_info = self.process_smart_image_logic(processed_body, excel_path, sheet_name, data_range)
                     if img_info and os.path.exists(img_info['path']):
                         smart_img_path = img_info['path']
             except Exception as e:
                 print(f"Preview Smart Image Error: {e}")

        # 渲染正文 (支持图片插入)
        self.preview_images = [] # 防止GC回收
        
        if "{SMART_IMAGE}" in processed_body:
            parts = processed_body.split("{SMART_IMAGE}")
            for i, part in enumerate(parts):
                body_text.insert(tk.END, part)
                if i < len(parts) - 1:
                    if smart_img_path and ImageTk:
                        try:
                            pil_img = Image.open(smart_img_path)
                            # 缩放以适应预览窗口
                            width_ratio = 750 / pil_img.width
                            if width_ratio < 1:
                                new_size = (int(pil_img.width * width_ratio), int(pil_img.height * width_ratio))
                                pil_img = pil_img.resize(new_size, Image.Resampling.LANCZOS)
                            
                            photo = ImageTk.PhotoImage(pil_img)
                            self.preview_images.append(photo) # Keep reference
                            body_text.image_create(tk.END, image=photo)
                            body_text.insert(tk.END, "\n")
                        except Exception as e:
                            body_text.insert(tk.END, f"\n[图片渲染错误: {str(e)}]\n")
                    else:
                        body_text.insert(tk.END, "\n[Excel智能截图占位]\n")
        else:
            body_text.insert(1.0, processed_body)

        body_text.config(state='disabled')
        
        # 底部信息
        info_frame = ttk.Frame(preview_win)
        info_frame.pack(fill=tk.X, padx=10, pady=5)
        
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ttk.Label(info_frame, text=f"预览生成时间: {now_str}", 
                 font=('TkDefaultFont', 8), foreground='gray').pack(side=tk.LEFT)
    
    def toggle_language(self):
        """切换语言"""
        self.current_language = 'en' if self.current_language == 'zh' else 'zh'
        
        # 更新窗口标题
        self.root.title(self.translations[self.current_language]['title'])
        
        # 更新状态栏
        if self.current_language == 'zh':
            self.status_var.set("就绪 | 快捷键: Ctrl+S=保存配置 | Ctrl+D=创建草稿 | Ctrl+P=预览 | F1=帮助")
            messagebox.showinfo("语言切换", "已切换到中文\n注意：部分界面需要重启程序才能完全更新")
        else:
            self.status_var.set("Ready | Shortcuts: Ctrl+S=Save | Ctrl+D=Draft | Ctrl+P=Preview | F1=Help")
            messagebox.showinfo("Language Switch", "Switched to English\nNote: Some UI elements require restart to fully update")
    
    def zoom_in(self):
        """放大界面"""
        if self.ui_scale < 2.0:  # 最大200%
            self.ui_scale += 0.1
            self.apply_zoom()
            self.save_ui_preferences()  # 保存设置
    
    def zoom_out(self):
        """缩小界面"""
        if self.ui_scale > 0.5:  # 最小50%
            self.ui_scale -= 0.1
            self.apply_zoom()
            self.save_ui_preferences()  # 保存设置
    
    def reset_zoom(self):
        """重置缩放"""
        self.ui_scale = 1.0
        self.apply_zoom()
        self.save_ui_preferences()  # 保存设置
    
    def mouse_zoom(self, event):
        """鼠标滚轮缩放"""
        if event.delta > 0:
            self.zoom_in()
        else:
            self.zoom_out()
    
    def apply_zoom(self):
        """应用缩放设置"""
        # 更新缩放标签
        zoom_percent = int(self.ui_scale * 100)
        self.zoom_label.config(text=f"{zoom_percent}%")
        
        # 计算字体大小
        base_font_size = 9
        new_font_size = int(base_font_size * self.ui_scale)
        
        # 更新默认字体
        default_font = ('TkDefaultFont', new_font_size)
        text_font = ('TkTextFont', new_font_size)
        fixed_font = ('TkFixedFont', new_font_size)
        
        try:
            self.root.option_add('*TCombobox*Listbox.font', default_font)
            self.root.option_add('*Font', default_font)
        except:
            pass
        
        # 更新所有文本框的字体
        for widget in self.root.winfo_children():
            self._update_widget_font(widget, new_font_size)
        
        # 更新窗口大小
        base_width = 1200
        base_height = 800
        new_width = int(base_width * self.ui_scale)
        new_height = int(base_height * self.ui_scale)
        
        # 限制最小和最大尺寸
        new_width = max(800, min(new_width, 1920))
        new_height = max(600, min(new_height, 1080))
        
        self.root.geometry(f"{new_width}x{new_height}")
        
        # 更新状态
        self.status_var.set(f"界面缩放: {zoom_percent}% | Ctrl+Plus/Minus 或 Ctrl+滚轮缩放 | Ctrl+0 重置")
    
    def _update_widget_font(self, widget, font_size):
        """递归更新控件字体"""
        try:
            # 更新当前控件
            widget_type = widget.winfo_class()
            
            if widget_type in ('TEntry', 'TCombobox', 'TLabel', 'TButton'):
                # ttk 控件通过 style 更新
                pass
            elif widget_type in ('Text', 'Entry', 'Label', 'Button', 'Listbox'):
                # tk 控件直接更新
                try:
                    current_font = widget.cget('font')
                    if current_font:
                        if isinstance(current_font, str):
                            # 字体名称
                            widget.config(font=(current_font, font_size))
                        elif isinstance(current_font, tuple):
                            # 字体元组
                            font_family = current_font[0] if len(current_font) > 0 else 'TkDefaultFont'
                            widget.config(font=(font_family, font_size))
                except:
                    pass
            
            # 递归处理子控件
            for child in widget.winfo_children():
                self._update_widget_font(child, font_size)
        except:
            pass
    
    def show_zoom_help(self):
        """显示缩放帮助"""
        help_text = """
═══════════════════════════════
    界面缩放功能
═══════════════════════════════

🔍 缩放方式:

1. 按钮控制:
   🔍+     放大界面
   🔍−     缩小界面
   
2. 键盘快捷键:
   Ctrl + Plus     放大
   Ctrl + Minus    缩小
   Ctrl + 0        重置为100%

3. 鼠标控制:
   Ctrl + 滚轮向上    放大
   Ctrl + 滚轮向下    缩小

📏 缩放范围:
   最小: 50%
   最大: 200%
   步进: 10%

💡 提示:
• 缩放会同时调整窗口大小和字体
• 适合不同屏幕分辨率和视力需求
• 设置会保持到程序关闭

═══════════════════════════════
        """
        messagebox.showinfo("缩放功能帮助", help_text)


def main():
    print("=== 程序启动 ===", flush=True)
    try:
        print("创建 Tk() 实例...", flush=True)
        root = tk.Tk()
        print("Tk() 创建成功", flush=True)
        
        print("创建 OutlookDraftManager...", flush=True)
        app = OutlookDraftManager(root)
        print("OutlookDraftManager 创建成功", flush=True)
        
        print("进入主循环...", flush=True)
        root.mainloop()
        print("主循环退出", flush=True)
    except Exception as e:
        import traceback
        error_msg = f"程序启动失败:\n{str(e)}\n\n详细错误:\n{traceback.format_exc()}"
        print(error_msg, flush=True)
        try:
            messagebox.showerror("致命错误", error_msg)
        except:
            pass
        input("按回车键退出...")


if __name__ == "__main__":
    print("=== __main__ 开始执行 ===", flush=True)
    main()
